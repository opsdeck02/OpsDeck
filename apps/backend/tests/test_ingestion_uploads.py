from __future__ import annotations

import io
import json
from collections.abc import Generator
from datetime import UTC, datetime
from decimal import Decimal
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import create_engine, func, select
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool

from app.api.dependencies import get_db
from app.db.base import Base
from app.main import app
from app.models import (
    ImportJobRecord,
    IngestionJob,
    LineStopIncident,
    Material,
    OperationalEvent,
    Plant,
    PlantMaterialThreshold,
    Role,
    Shipment,
    StockSnapshot,
    Supplier,
    Tenant,
    TenantMembership,
    UploadedFile,
    User,
)
from app.models.enums import OperationalEventFreshnessStatus, OperationalEventType
from app.modules.auth.constants import LOGISTICS_USER, TENANT_ADMIN
from app.modules.auth.security import hash_password
from app.modules.line_stops.service import build_historical_validation_report
from app.modules.stock.service import calculate_stock_cover_detail
from app.schemas.context import RequestContext

FIXTURE_DIR = Path(__file__).resolve().parent / "fixtures" / "ingestion"


@pytest.fixture()
def client_and_session() -> Generator[tuple[TestClient, sessionmaker[Session]], None, None]:
    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    TestingSessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    Base.metadata.create_all(bind=engine)

    with TestingSessionLocal() as db:
        seed_ingestion_test_data(db)

    def override_get_db() -> Generator[Session, None, None]:
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[get_db] = override_get_db
    try:
        yield TestClient(app), TestingSessionLocal
    finally:
        app.dependency_overrides.clear()
        Base.metadata.drop_all(bind=engine)


def seed_ingestion_test_data(db: Session) -> None:
    tenant = Tenant(name="Tenant A", slug="tenant-a")
    role = Role(name=LOGISTICS_USER, description="Logistics")
    admin_role = Role(name=TENANT_ADMIN, description="Tenant admin")
    db.add_all([tenant, role, admin_role])
    db.flush()
    user = User(
        email="logistics@test.local",
        full_name="Logistics User",
        password_hash=hash_password("TestOnlyCredential1!"),
        is_active=True,
    )
    admin = User(
        email="admin@test.local",
        full_name="Admin User",
        password_hash=hash_password("TestOnlyCredential1!"),
        is_active=True,
    )
    plant = Plant(tenant_id=tenant.id, code="JAM", name="Demo Plant A", location="Jharkhand")
    material = Material(
        tenant_id=tenant.id,
        code="COKING_COAL",
        name="Coking coal",
        category="coal",
        uom="MT",
    )
    db.add_all([user, admin, plant, material])
    db.flush()
    db.add_all([
        TenantMembership(
            tenant_id=tenant.id,
            user_id=user.id,
            role_id=role.id,
            is_active=True,
        ),
        TenantMembership(
            tenant_id=tenant.id,
            user_id=admin.id,
            role_id=admin_role.id,
            is_active=True,
        ),
    ])
    db.commit()


def login(client: TestClient) -> dict[str, str]:
    response = client.post(
        "/api/v1/auth/login",
        json={"email": "logistics@test.local", "password": "TestOnlyCredential1!"},
    )
    assert response.status_code == 200, response.json()
    token = response.json()["access_token"]
    return {"Authorization": f"Bearer {token}", "X-Tenant-Slug": "tenant-a"}


def admin_login(client: TestClient) -> dict[str, str]:
    response = client.post(
        "/api/v1/auth/login",
        json={"email": "admin@test.local", "password": "TestOnlyCredential1!"},
    )
    assert response.status_code == 200, response.json()
    token = response.json()["access_token"]
    return {"Authorization": f"Bearer {token}", "X-Tenant-Slug": "tenant-a"}


def enable_demo_tenant(SessionLocal: sessionmaker[Session]) -> None:
    with SessionLocal() as db:
        tenant = db.scalar(select(Tenant).where(Tenant.slug == "tenant-a"))
        assert tenant is not None
        tenant.is_demo_tenant = True
        db.commit()


def upload_csv(client: TestClient, headers: dict[str, str], file_type: str, csv_body: str):
    return client.post(
        "/api/v1/ingestion/uploads",
        headers=headers,
        data={"file_type": file_type},
        files={"file": ("upload.csv", csv_body.encode(), "text/csv")},
    )


def upload_xlsx(client: TestClient, headers: dict[str, str], file_type: str, content: bytes):
    return client.post(
        "/api/v1/ingestion/uploads",
        headers=headers,
        data={"file_type": file_type},
        files={
            "file": (
                "stock_snapshot.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )


def upload_xlsm(client: TestClient, headers: dict[str, str], file_type: str, content: bytes):
    return client.post(
        "/api/v1/ingestion/uploads",
        headers=headers,
        data={"file_type": file_type},
        files={
            "file": (
                "shipment_report.xlsm",
                content,
                "application/vnd.ms-excel.sheet.macroEnabled.12",
            )
        },
    )


def upload_workbook(
    client: TestClient,
    headers: dict[str, str],
    content: bytes,
    sheet_configs: list[dict[str, object]],
):
    return client.post(
        "/api/v1/ingestion/workbook-upload",
        headers=headers,
        data={"sheet_configs": json.dumps(sheet_configs)},
        files={
            "file": (
                "operations.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )


def test_valid_shipment_upload(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, SessionLocal = client_and_session
    response = upload_csv(client, login(client), "shipment", shipment_csv("SHP-001"))

    assert response.status_code == 200, response.json()
    body = response.json()
    assert body["rows_received"] == 1
    assert body["rows_accepted"] == 1
    assert body["summary_counts"]["created"] == 1
    assert body["operational_summary"]["shipments_detected"] == ["SHP-001"]
    assert body["operational_summary"]["plants_detected"] == ["JAM"]
    assert body["operational_summary"]["materials_detected"] == ["COKING_COAL"]
    assert body["operational_summary"]["refreshed_operational_visibility"] is True
    assert body["operational_summary"]["next_recommended_action"]
    assert body["operational_summary"]["supplier_references_total"] == 1
    assert body["operational_summary"]["supplier_references_linked"] == 1
    assert body["operational_summary"]["supplier_references_unlinked"] == 0
    assert body["operational_summary"]["onboarding_completeness_score"] == 100
    assert body["operational_summary"]["supplier_reliability_impact"] is None

    with SessionLocal() as db:
        assert db.scalar(select(func.count()).select_from(Shipment)) == 1
        supplier = db.scalar(select(Supplier).where(Supplier.name == "Supplier A"))
        assert supplier is not None
        shipment = db.scalar(select(Shipment).where(Shipment.shipment_id == "SHP-001"))
        assert shipment is not None
        assert shipment.supplier_id == supplier.id
        event = db.scalar(select(OperationalEvent))
        assert event is not None
        assert event.tenant_id == 1
        assert event.event_type == OperationalEventType.SHIPMENT_MILESTONE_UPDATED
        assert event.event_category == "shipment"
        assert event.shipment_reference == "SHP-001"
        assert event.source_type == "manual_upload"
        assert event.source_id is not None
        assert event.confidence_score is not None
        assert event.confidence_score > 0
        assert event.metadata_json is not None
        assert event.metadata_json["confidence"]["score"] == float(event.confidence_score)
        assert "source_reliability" in event.metadata_json["confidence"]["factors"]


def test_shipment_upload_deduplicates_detected_supplier_names(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, SessionLocal = client_and_session
    csv_body = "\n".join(
        [
            "shipment_id,plant_code,material_code,supplier_name,quantity_mt,"
            "planned_eta,current_eta,current_state,latest_update_at",
            "SHP-SUP-1,JAM,COKING_COAL,Supplier A,10,"
            "2026-04-20T08:00:00Z,2026-04-21T08:00:00Z,"
            "in_transit,2026-04-15T09:00:00Z",
            "SHP-SUP-2,JAM,COKING_COAL,supplier a,20,"
            "2026-04-22T08:00:00Z,2026-04-23T08:00:00Z,"
            "in_transit,2026-04-16T09:00:00Z",
        ]
    )

    response = upload_csv(client, login(client), "shipment", csv_body)

    assert response.status_code == 200, response.json()
    with SessionLocal() as db:
        suppliers = list(db.scalars(select(Supplier)))
        assert len(suppliers) == 1
        assert suppliers[0].name == "Supplier A"
        assert suppliers[0].code == "SUPPLIER_A"
        assert db.scalar(select(func.count()).select_from(Shipment)) == 2
        assert (
            db.scalar(
                select(func.count())
                .select_from(Shipment)
                .where(Shipment.supplier_id == suppliers[0].id)
            )
            == 2
        )


def test_duplicate_shipment_id_is_reported(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, _ = client_and_session
    csv_body = "\n".join(
        [
            "shipment_id,plant_code,material_code,supplier_name,quantity_mt,"
            "planned_eta,current_eta,current_state,latest_update_at",
            "SHP-DUP,JAM,COKING_COAL,Supplier A,10,"
            "2026-04-20T08:00:00Z,2026-04-21T08:00:00Z,"
            "in_transit,2026-04-15T09:00:00Z",
            "SHP-DUP,JAM,COKING_COAL,Supplier A,12,"
            "2026-04-22T08:00:00Z,2026-04-23T08:00:00Z,"
            "in_transit,2026-04-16T09:00:00Z",
        ]
    )

    response = upload_csv(client, login(client), "shipment", csv_body)

    assert response.status_code == 200, response.json()
    summary = response.json()["operational_summary"]
    assert summary["duplicate_rows_detected"] == 1
    assert any("duplicate row" in warning.lower() for warning in summary["warnings"])


def test_duplicate_stock_snapshot_row_is_reported(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, _ = client_and_session
    csv_body = "\n".join(
        [
            "plant_code,material_code,on_hand_mt,quality_held_mt,available_to_consume_mt,daily_consumption_mt,snapshot_time",
            "JAM,COKING_COAL,10000,1000,9000,500,2026-04-15T08:00:00Z",
            "JAM,COKING_COAL,10500,1000,9500,500,2026-04-15T08:00:00Z",
        ]
    )

    response = upload_csv(client, login(client), "stock", csv_body)

    assert response.status_code == 200, response.json()
    assert response.json()["operational_summary"]["duplicate_rows_detected"] == 1


def test_duplicate_threshold_row_is_reported(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, _ = client_and_session
    csv_body = "\n".join(
        [
            "plant_code,material_code,threshold_days,warning_days",
            "JAM,COKING_COAL,7,10",
            "JAM,COKING_COAL,8,11",
        ]
    )

    response = upload_csv(client, login(client), "threshold", csv_body)

    assert response.status_code == 200, response.json()
    assert response.json()["operational_summary"]["duplicate_rows_detected"] == 1


def test_auto_created_material_is_reported_in_summary(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, _ = client_and_session
    csv_body = "\n".join(
        [
            "plant_code,material_code,on_hand_mt,quality_held_mt,available_to_consume_mt,daily_consumption_mt,snapshot_time",
            "JAM,PCI_COAL,10000,1000,9000,500,2026-04-15T08:00:00Z",
        ]
    )

    response = upload_csv(client, login(client), "stock", csv_body)

    assert response.status_code == 200, response.json()
    summary = response.json()["operational_summary"]
    assert summary["new_materials_created"] == ["PCI_COAL"]
    assert any("created new master data" in warning for warning in summary["warnings"])


def test_seeded_pilot_template_stock_upload_does_not_report_new_master_data(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, _ = client_and_session
    csv_body = "\n".join(
        [
            "plant_code,material_code,on_hand_mt,quality_held_mt,available_to_consume_mt,daily_consumption_mt,snapshot_time",
            "JAM,COKING_COAL,180000,10000,170000,24000,2026-04-15T08:00:00Z",
        ]
    )

    response = upload_csv(client, login(client), "stock", csv_body)

    assert response.status_code == 200, response.json()
    summary = response.json()["operational_summary"]
    assert summary["new_plants_created"] == []
    assert summary["new_materials_created"] == []
    assert not any("created new master data" in warning for warning in summary["warnings"])


def test_download_templates_only_include_pilot_critical_columns(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, _ = client_and_session
    headers = login(client)

    shipment_response = client.get("/api/v1/ingestion/templates/shipment", headers=headers)
    consumption_response = client.get("/api/v1/ingestion/templates/consumption", headers=headers)

    assert shipment_response.status_code == 200, shipment_response.text
    shipment_header = shipment_response.text.splitlines()[0]
    assert shipment_header == (
        "shipment_id,plant_code,material_code,supplier_name,quantity_mt,current_eta,"
        "current_state,latest_update_at,eta_confidence"
    )
    assert "vessel_name" not in shipment_header
    assert "imo_number" not in shipment_header
    assert "origin_port" not in shipment_header
    assert consumption_response.status_code == 200, consumption_response.text
    assert consumption_response.text.splitlines()[0] == (
        "plant_code,material_code,daily_consumption_mt,snapshot_time"
    )


def test_missing_eta_gives_human_warning(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, _ = client_and_session
    csv_body = "\n".join(
        [
            "shipment_id,plant_code,material_code,supplier_name,quantity_mt,current_state,latest_update_at",
            "SHP-NO-ETA,JAM,COKING_COAL,Supplier A,10,in_transit,2026-04-15T09:00:00Z",
        ]
    )

    response = upload_csv(client, login(client), "shipment", csv_body)

    assert response.status_code == 400
    body = response.json()["detail"]
    summary = body["operational_summary"]
    assert summary["missing_eta_count"] == 1
    assert any(
        "ETA is required because OpsDeck cannot judge" in warning
        for warning in summary["warnings"]
    )


def test_missing_consumption_gives_setup_warning(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, _ = client_and_session
    csv_body = "\n".join(
        [
            "plant_code,material_code,on_hand_mt,quality_held_mt,available_to_consume_mt,daily_consumption_mt,snapshot_time",
            "JAM,COKING_COAL,10000,1000,9000,,2026-04-15T08:00:00Z",
        ]
    )

    response = upload_csv(client, login(client), "stock", csv_body)

    assert response.status_code == 400
    summary = response.json()["detail"]["operational_summary"]
    assert summary["missing_consumption_count"] == 1
    assert "Daily consumption is required" in summary["warnings"][0]


def test_missing_threshold_gives_setup_warning(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, _ = client_and_session
    csv_body = "\n".join(
        [
            "plant_code,material_code,threshold_days,warning_days",
            "JAM,COKING_COAL,,10",
        ]
    )

    response = upload_csv(client, login(client), "threshold", csv_body)

    assert response.status_code == 400
    summary = response.json()["detail"]["operational_summary"]
    assert summary["missing_threshold_count"] == 1
    assert "Thresholds are required" in summary["warnings"][0]


def test_indian_style_headers_map_correctly(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, _ = client_and_session
    csv_body = "\n".join(
        [
            "Rake No,Plant,Material,Vendor,Quantity,ETA at Plant,Status,Last Updated",
            "R-101,JAM,COKING_COAL,Supplier A,100,2026-04-21T08:00:00Z,"
            "in_transit,2026-04-15T09:00:00Z",
        ]
    )

    response = upload_csv(client, login(client), "shipment", csv_body)

    assert response.status_code == 200, response.json()
    body = response.json()
    assert body["rows_accepted"] == 1
    assert body["operational_summary"]["shipments_detected"] == ["R-101"]


def test_shipment_upload_missing_supplier_is_rejected_without_broken_supplier(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, SessionLocal = client_and_session
    csv_body = "\n".join(
        [
            "shipment_id,plant_code,material_code,supplier_name,quantity_mt,"
            "planned_eta,current_eta,current_state,latest_update_at",
            "SHP-MISSING-SUP,JAM,COKING_COAL,,10,"
            "2026-04-20T08:00:00Z,2026-04-21T08:00:00Z,"
            "in_transit,2026-04-15T09:00:00Z",
        ]
    )

    response = upload_csv(client, login(client), "shipment", csv_body)

    assert response.status_code == 400, response.json()
    body = response.json()["detail"]
    assert body["rows_accepted"] == 0
    assert body["rows_rejected"] == 1
    field_error = body["validation_errors"][0]["field_errors"][0]
    assert field_error["field"] == "supplier_name"
    assert "Reliability source is missing" in field_error["reason"]
    with SessionLocal() as db:
        assert db.scalar(select(func.count()).select_from(Supplier)) == 0
        assert db.scalar(select(func.count()).select_from(Shipment)) == 0


def test_valid_stock_upload(client_and_session: tuple[TestClient, sessionmaker[Session]]) -> None:
    client, SessionLocal = client_and_session
    snapshot_time = datetime.now(UTC).isoformat()
    response = upload_csv(
        client,
        login(client),
        "stock",
        "\n".join(
            [
                "plant_code,material_code,on_hand_mt,quality_held_mt,available_to_consume_mt,daily_consumption_mt,snapshot_time",
                f"JAM,COKING_COAL,10000,1000,9000,500,{snapshot_time}",
            ]
        ),
    )

    assert response.status_code == 200
    assert response.json()["summary_counts"]["created"] == 1
    with SessionLocal() as db:
        assert db.scalar(select(func.count()).select_from(StockSnapshot)) == 1
        event = db.scalar(select(OperationalEvent))
        assert event is not None
        assert event.tenant_id == 1
        assert event.event_type == OperationalEventType.INVENTORY_STOCK_UPDATED
        assert event.event_category == "inventory"
        assert event.plant_reference == "JAM"
        assert event.material_reference == "COKING_COAL"
        assert str(event.quantity_value) == "9000.000"
        assert event.source_id is not None
        assert event.confidence_score is not None
        assert event.metadata_json is not None
        assert event.metadata_json["confidence"]["factors"]["completeness"] == 100
        assert event.metadata_json["confidence"]["reasons"]
        assert event.freshness_status == OperationalEventFreshnessStatus.FRESH
        assert event.metadata_json["freshness"]["status"] == "fresh"
        assert event.metadata_json["confidence"]


def test_shipment_eta_update_creates_operational_event(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, SessionLocal = client_and_session
    headers = login(client)

    first_response = upload_csv(client, headers, "shipment", shipment_csv("SHP-ETA-EVENT"))
    csv_header = (
        "shipment_id,plant_code,material_code,supplier_name,quantity_mt,"
        "planned_eta,current_eta,current_state,latest_update_at"
    )
    csv_row = (
        "SHP-ETA-EVENT,JAM,COKING_COAL,Supplier A,74000,"
        "2026-04-20T08:00:00Z,2026-04-23T08:00:00Z,"
        "in_transit,2026-04-16T09:00:00Z"
    )
    update_response = upload_csv(
        client,
        headers,
        "shipment",
        "\n".join([csv_header, csv_row]),
    )

    assert first_response.status_code == 200
    assert update_response.status_code == 200, update_response.json()
    assert update_response.json()["summary_counts"]["updated"] == 1
    with SessionLocal() as db:
        event = db.scalar(
            select(OperationalEvent)
            .where(OperationalEvent.event_type == OperationalEventType.SHIPMENT_ETA_CHANGED)
            .order_by(OperationalEvent.id.desc())
        )
        assert event is not None
        assert event.shipment_reference == "SHP-ETA-EVENT"
        assert event.previous_value is not None
        assert event.previous_value["current_eta"] == "2026-04-21T08:00:00+00:00"
        assert event.new_value is not None
        assert event.new_value["current_eta"] == "2026-04-23T08:00:00+00:00"
        assert event.confidence_score is not None
        assert event.metadata_json is not None
        assert "confidence" in event.metadata_json
        assert "freshness" in event.metadata_json


def test_stock_xlsx_detects_headers_on_row_three(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, SessionLocal = client_and_session

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Stock snapshot export"])
    sheet.append(["Generated from customer workbook"])
    sheet.append(
        [
            "last_updated_at",
            "plant_code",
            "material_code",
            "material_name",
            "current_stock_tons",
            "available_unrestricted_tons",
            "blocked_stock_tons",
            "in_transit_open_tons",
            "daily_consumption_tons",
            "days_to_line_stop",
            "risk_status",
            "next_inbound_eta_days",
        ]
    )
    sheet.append(
        [
            "2026-04-15T08:00:00Z",
            "JAM",
            "COKING_COAL",
            "Coking coal",
            10000,
            9000,
            1000,
            2500,
            500,
            18,
            "safe",
            4,
        ]
    )
    output = io.BytesIO()
    workbook.save(output)

    response = upload_xlsx(client, login(client), "stock", output.getvalue())

    assert response.status_code == 200, response.json()
    assert response.json()["rows_accepted"] == 1
    with SessionLocal() as db:
        snapshot = db.scalar(select(StockSnapshot))
        assert snapshot is not None
        assert str(snapshot.on_hand_mt) == "10000.000"
        assert str(snapshot.available_to_consume_mt) == "9000.000"
        assert str(snapshot.quality_held_mt) == "1000.000"


def test_shipment_xlsm_upload_is_supported(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, SessionLocal = client_and_session
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "shipment_id",
            "plant_code",
            "material_code",
            "supplier_name",
            "quantity_mt",
            "planned_eta",
            "current_eta",
            "current_state",
            "latest_update_at",
        ]
    )
    sheet.append(
        [
            "SHP-XLSM-1",
            "JAM",
            "COKING_COAL",
            "Supplier A",
            100,
            "2026-05-10T00:00:00Z",
            "2026-05-10T00:00:00Z",
            "in_transit",
            "2026-05-06T08:00:00Z",
        ]
    )
    output = io.BytesIO()
    workbook.save(output)

    response = upload_xlsm(client, login(client), "shipment", output.getvalue())

    assert response.status_code == 200, response.json()
    assert response.json()["rows_accepted"] == 1
    with SessionLocal() as db:
        shipment = db.scalar(select(Shipment).where(Shipment.shipment_id == "SHP-XLSM-1"))
        assert shipment is not None


def test_shipment_upload_creates_unknown_plant(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, SessionLocal = client_and_session
    response = upload_csv(
        client,
        login(client),
        "shipment",
        "\n".join(
            [
                "shipment_id,plant_code,material_code,supplier_name,quantity_mt,"
                "planned_eta,current_eta,current_state,latest_update_at",
                "SHP-NEW-PLANT,DEMO_PLANT_A,COKING_COAL,Supplier A,100,"
                "2026-05-10T00:00:00Z,2026-05-10T00:00:00Z,in_transit,"
                "2026-05-06T08:00:00Z",
            ]
        ),
    )

    assert response.status_code == 200, response.json()
    assert response.json()["rows_accepted"] == 1
    with SessionLocal() as db:
        plant = db.scalar(select(Plant).where(Plant.code == "DEMO_PLANT_A"))
        shipment = db.scalar(select(Shipment).where(Shipment.shipment_id == "SHP-NEW-PLANT"))
        assert plant is not None
        assert plant.name == "DEMO_PLANT_A"
        assert shipment is not None
        assert shipment.plant_id == plant.id


def test_invalid_stock_upload_rejects_zero_daily_consumption(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, SessionLocal = client_and_session
    response = upload_csv(
        client,
        login(client),
        "stock",
        "\n".join(
            [
                "plant_code,material_code,on_hand_mt,quality_held_mt,available_to_consume_mt,daily_consumption_mt,snapshot_time",
                "JAM,COKING_COAL,10000,1000,9000,0,2026-04-15T08:00:00Z",
            ]
        ),
    )

    assert response.status_code == 400
    assert "Daily consumption MT must be greater than zero" in str(response.json())
    with SessionLocal() as db:
        assert db.scalar(select(func.count()).select_from(StockSnapshot)) == 0


def test_missing_required_column_blocks_ingestion(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, SessionLocal = client_and_session
    response = upload_csv(
        client,
        login(client),
        "stock",
        "\n".join(
            [
                "plant_code,material_code,on_hand_mt,quality_held_mt,daily_consumption_mt,snapshot_time",
                "JAM,COKING_COAL,10000,1000,500,14/05/2026",
            ]
        ),
    )

    assert response.status_code == 400
    assert "Missing required mapping: Available to consume MT" in str(response.json())
    with SessionLocal() as db:
        assert db.scalar(select(func.count()).select_from(StockSnapshot)) == 0


def test_header_alias_case_space_matching_works(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, SessionLocal = client_and_session
    response = upload_csv(
        client,
        login(client),
        "stock",
        "\n".join(
            [
                " Plant Code , MATERIAL CODE , Current Stock Tons , Blocked Stock Tons , "
                "Available Unrestricted Tons , Daily Consumption Tons , Last Updated At ",
                "JAM,COKING_COAL,10000,1000,9000,500,14/05/2026",
            ]
        ),
    )

    assert response.status_code == 200, response.json()
    assert response.json()["rows_accepted"] == 1
    with SessionLocal() as db:
        assert db.scalar(select(func.count()).select_from(StockSnapshot)) == 1


def test_rejected_row_includes_field_reason_and_suggested_fix(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, _ = client_and_session
    response = upload_csv(
        client,
        login(client),
        "stock",
        "\n".join(
            [
                "plant_code,material_code,on_hand_mt,quality_held_mt,available_to_consume_mt,daily_consumption_mt,snapshot_time",
                "JAM,COKING_COAL,10000,1000,9000,abc,14/05/2026",
            ]
        ),
    )

    assert response.status_code == 400
    body = response.json()["detail"]
    assert body["rows_received"] == 1
    assert body["rows_accepted"] == 0
    assert body["rows_rejected"] == 1
    row_error = body["validation_errors"][0]
    assert row_error["row_number"] == 2
    assert row_error["field_errors"][0]["field"] == "daily_consumption_mt"
    assert "could not be interpreted" in row_error["field_errors"][0]["reason"]
    assert row_error["field_errors"][0]["suggested_fix"]
    assert body["top_rejection_reasons"][0]["count"] == 1
    assert body["operational_summary"]["rows_rejected"] == 1
    assert body["operational_summary"]["next_recommended_action"].startswith("Fix rejected rows")


def test_common_date_formats_and_numeric_strings_with_commas_parse(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, SessionLocal = client_and_session
    response = upload_csv(
        client,
        login(client),
        "stock",
        "\n".join(
            [
                "plant_code,material_code,on_hand_mt,quality_held_mt,available_to_consume_mt,daily_consumption_mt,snapshot_time",
                'JAM,COKING_COAL,"10,000 MT","1,000 tonnes","9,000","500",14-05-2026 09:30',
            ]
        ),
    )

    assert response.status_code == 200, response.json()
    with SessionLocal() as db:
        snapshot = db.scalar(select(StockSnapshot))
        assert snapshot is not None
        assert str(snapshot.on_hand_mt) == "10000.000"
        assert snapshot.snapshot_time.isoformat() == "2026-05-14T09:30:00"


def test_workbook_with_inventory_and_inbound_sheets_ingests_successfully(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, SessionLocal = client_and_session
    workbook = Workbook()
    stock_sheet = workbook.active
    stock_sheet.title = "Current Stock"
    stock_sheet.append(["Inventory workbook"])
    stock_sheet.append(
        [
            "plant_code",
            "material_code",
            "on_hand_mt",
            "quality_held_mt",
            "available_to_consume_mt",
            "daily_consumption_mt",
            "snapshot_time",
        ]
    )
    stock_sheet.append(["JAM", "COKING_COAL", "10,000 MT", 1000, 9000, 500, "15/05/2026"])
    inbound_sheet = workbook.create_sheet("Inbound ETA")
    inbound_sheet.append(
        [
            "shipment_id",
            "plant_code",
            "material_code",
            "supplier_name",
            "quantity_mt",
            "planned_eta",
            "current_eta",
            "current_state",
            "latest_update_at",
        ]
    )
    inbound_sheet.append(
        [
            "WB-IN-1",
            "JAM",
            "COKING_COAL",
            "Supplier A",
            74000,
            "15-05-2026",
            "17-05-2026",
            "in_transit",
            "15-05-2026 09:00",
        ]
    )
    workbook.create_sheet("Notes").append(["owner notes only"])
    workbook.create_sheet("Empty Sheet")
    output = io.BytesIO()
    workbook.save(output)

    headers = login(client)
    preview_response = client.post(
        "/api/v1/ingestion/workbook-preview",
        headers=headers,
        files={
            "file": (
                "operations.xlsx",
                output.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert preview_response.status_code == 200, preview_response.json()
    preview = preview_response.json()
    suggestions = {sheet["sheet_name"]: sheet["suggested_file_type"] for sheet in preview["sheets"]}
    assert suggestions["Current Stock"] == "stock"
    assert suggestions["Inbound ETA"] == "shipment"
    assert "Empty Sheet" in preview["ignored_empty_sheets"]

    upload_response = upload_workbook(
        client,
        headers,
        output.getvalue(),
        [
            {"sheet_name": "Current Stock", "file_type": "stock", "mapping_overrides": {}},
            {"sheet_name": "Inbound ETA", "file_type": "shipment", "mapping_overrides": {}},
            {"sheet_name": "Notes", "file_type": "ignore", "mapping_overrides": {}},
        ],
    )

    assert upload_response.status_code == 200, upload_response.json()
    body = upload_response.json()
    assert body["file_type"] == "workbook"
    assert body["rows_accepted"] == 2
    assert body["operational_summary"]["plants_detected"] == ["JAM"]
    assert body["operational_summary"]["materials_detected"] == ["COKING_COAL"]
    assert body["operational_summary"]["shipments_detected"] == ["WB-IN-1"]
    assert body["operational_summary"]["refreshed_operational_visibility"] is True
    assert {sheet["sheet_name"] for sheet in body["sheet_results"]} == {
        "Current Stock",
        "Inbound ETA",
    }
    assert "Notes" in body["ignored_sheets"]
    with SessionLocal() as db:
        assert db.scalar(select(func.count()).select_from(StockSnapshot)) == 1
        assert db.scalar(select(func.count()).select_from(Shipment)) == 1
        import_records = list(db.scalars(select(ImportJobRecord)))
        assert import_records
        assert all(record.created_at is not None for record in import_records)
        assert all(record.updated_at is not None for record in import_records)


def test_workbook_missing_required_mapping_blocks_only_affected_sheet(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, SessionLocal = client_and_session
    workbook = Workbook()
    stock_sheet = workbook.active
    stock_sheet.title = "Inventory"
    stock_sheet.append(
        [
            "plant_code",
            "material_code",
            "on_hand_mt",
            "quality_held_mt",
            "daily_consumption_mt",
            "snapshot_time",
        ]
    )
    stock_sheet.append(["JAM", "COKING_COAL", 10000, 1000, 500, "15/05/2026"])
    inbound_sheet = workbook.create_sheet("Inbound")
    inbound_sheet.append(
        [
            "shipment_id",
            "plant_code",
            "material_code",
            "supplier_name",
            "quantity_mt",
            "planned_eta",
            "current_eta",
            "current_state",
            "latest_update_at",
        ]
    )
    inbound_sheet.append(
        [
            "WB-IN-2",
            "JAM",
            "COKING_COAL",
            "Supplier A",
            100,
            "15/05/2026",
            "16/05/2026",
            "in_transit",
            "15/05/2026",
        ]
    )
    output = io.BytesIO()
    workbook.save(output)

    response = upload_workbook(
        client,
        login(client),
        output.getvalue(),
        [
            {"sheet_name": "Inventory", "file_type": "stock", "mapping_overrides": {}},
            {"sheet_name": "Inbound", "file_type": "shipment", "mapping_overrides": {}},
        ],
    )

    assert response.status_code == 200, response.json()
    body = response.json()
    results = {sheet["sheet_name"]: sheet for sheet in body["sheet_results"]}
    assert results["Inventory"]["status"] == "failed"
    assert "Available to consume MT" in results["Inventory"]["blocking_errors"][0]
    assert results["Inbound"]["rows_accepted"] == 1
    with SessionLocal() as db:
        assert db.scalar(select(func.count()).select_from(StockSnapshot)) == 0
        assert db.scalar(select(func.count()).select_from(Shipment)) == 1


def test_workbook_consumption_sheet_updates_matching_inventory_snapshot(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, SessionLocal = client_and_session
    workbook = Workbook()
    inventory = workbook.active
    inventory.title = "Inventory"
    inventory.append(
        [
            "plant_code",
            "material_code",
            "on_hand_mt",
            "quality_held_mt",
            "available_to_consume_mt",
            "daily_consumption_mt",
            "snapshot_time",
        ]
    )
    inventory.append(["JAM", "COKING_COAL", 10000, 1000, 9000, 500, "15/05/2026"])
    consumption = workbook.create_sheet("Consumption")
    consumption.append(["plant_code", "material_code", "daily_consumption_mt", "snapshot_time"])
    consumption.append(["JAM", "COKING_COAL", 650, "15/05/2026"])
    output = io.BytesIO()
    workbook.save(output)

    response = upload_workbook(
        client,
        login(client),
        output.getvalue(),
        [
            {"sheet_name": "Inventory", "file_type": "stock", "mapping_overrides": {}},
            {"sheet_name": "Consumption", "file_type": "consumption", "mapping_overrides": {}},
        ],
    )

    assert response.status_code == 200, response.json()
    assert response.json()["rows_accepted"] == 2
    with SessionLocal() as db:
        snapshot = db.scalar(select(StockSnapshot))
        assert snapshot is not None
        assert str(snapshot.daily_consumption_mt) == "650.000"


def test_upload_history_records_counts_and_rejection_summary(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, _ = client_and_session
    headers = login(client)
    response = upload_csv(
        client,
        headers,
        "stock",
        "\n".join(
            [
                "plant_code,material_code,on_hand_mt,quality_held_mt,available_to_consume_mt,daily_consumption_mt,snapshot_time",
                "JAM,COKING_COAL,10000,1000,9000,abc,14/05/2026",
            ]
        ),
    )
    assert response.status_code == 400

    history_response = client.get("/api/v1/ingestion/jobs", headers=headers)
    assert history_response.status_code == 200
    job = history_response.json()[0]
    assert job["file_name"] == "upload.csv"
    assert job["source_type"] == "stock"
    assert job["rows_received"] == 1
    assert job["rows_accepted"] == 0
    assert job["rows_rejected"] == 1
    assert "Daily consumption MT could not be interpreted" in job["top_rejection_summary"]
    assert job["refreshed_operational_visibility"] is False


def test_import_job_detail_returns_summary_records_and_row_errors(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, _ = client_and_session
    headers = login(client)
    response = upload_csv(
        client,
        headers,
        "stock",
        "\n".join(
            [
                "plant_code,material_code,on_hand_mt,quality_held_mt,available_to_consume_mt,daily_consumption_mt,snapshot_time",
                "JAM,COKING_COAL,10000,1000,9000,abc,14/05/2026",
            ]
        ),
    )
    assert response.status_code == 400
    job_id = response.json()["detail"]["ingestion_job_id"]

    detail_response = client.get(f"/api/v1/ingestion/jobs/{job_id}", headers=headers)

    assert detail_response.status_code == 200, detail_response.json()
    detail = detail_response.json()
    assert detail["import_job_id"] == job_id
    assert detail["status"] == "failed"
    assert detail["operational_summary"]["rows_rejected"] == 1
    assert detail["row_level_errors"][0]["row_number"] == 2
    assert detail["record_references"] == []


def test_import_job_rollback_deletes_only_records_created_by_that_job(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, SessionLocal = client_and_session
    headers = login(client)
    first = upload_csv(client, headers, "shipment", shipment_csv("SHP-ROLLBACK-1"))
    second = upload_csv(client, headers, "shipment", shipment_csv("SHP-ROLLBACK-2"))
    assert first.status_code == 200
    assert second.status_code == 200

    rollback_response = client.post(
        f"/api/v1/ingestion/jobs/{first.json()['ingestion_job_id']}/rollback",
        headers=headers,
    )

    assert rollback_response.status_code == 200, rollback_response.json()
    body = rollback_response.json()
    assert body["rollback_status"] == "rolled_back"
    assert body["records_deleted"] == 1
    with SessionLocal() as db:
        assert db.scalar(select(Shipment).where(Shipment.shipment_id == "SHP-ROLLBACK-1")) is None
        assert (
            db.scalar(select(Shipment).where(Shipment.shipment_id == "SHP-ROLLBACK-2")) is not None
        )


def test_import_job_rollback_preserves_updated_preexisting_records(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, SessionLocal = client_and_session
    headers = login(client)
    first = upload_csv(client, headers, "shipment", shipment_csv("SHP-UPDATE-PRESERVE"))
    assert first.status_code == 200
    update = upload_csv(
        client,
        headers,
        "shipment",
        "\n".join(
            [
                "shipment_id,plant_code,material_code,supplier_name,quantity_mt,"
                "planned_eta,current_eta,current_state,latest_update_at",
                "SHP-UPDATE-PRESERVE,JAM,COKING_COAL,Supplier A,74000,"
                "2026-04-20T08:00:00Z,2026-04-24T08:00:00Z,"
                "in_transit,2026-04-16T09:00:00Z",
            ]
        ),
    )
    assert update.status_code == 200

    rollback_response = client.post(
        f"/api/v1/ingestion/jobs/{update.json()['ingestion_job_id']}/rollback",
        headers=headers,
    )

    assert rollback_response.status_code == 200
    body = rollback_response.json()
    assert body["records_deleted"] == 0
    assert body["records_preserved"] == 1
    assert "Exact update rollback is not available" in body["warnings"][0]
    with SessionLocal() as db:
        shipment = db.scalar(select(Shipment).where(Shipment.shipment_id == "SHP-UPDATE-PRESERVE"))
        assert shipment is not None
        assert shipment.current_eta.isoformat() == "2026-04-24T08:00:00"


def test_import_job_reprocess_uses_stored_file_without_duplicate_explosion(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, SessionLocal = client_and_session
    headers = login(client)
    response = upload_csv(client, headers, "shipment", shipment_csv("SHP-REPROCESS"))
    assert response.status_code == 200

    reprocess_response = client.post(
        f"/api/v1/ingestion/jobs/{response.json()['ingestion_job_id']}/reprocess",
        headers=headers,
    )

    assert reprocess_response.status_code == 200, reprocess_response.json()
    assert reprocess_response.json()["summary_counts"]["unchanged"] == 1
    with SessionLocal() as db:
        assert (
            db.scalar(
                select(func.count())
                .select_from(Shipment)
                .where(Shipment.shipment_id == "SHP-REPROCESS")
            )
            == 1
        )


def test_import_job_rollback_is_tenant_scoped(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, SessionLocal = client_and_session
    tenant_a_headers = login(client)
    upload_response = upload_csv(
        client,
        tenant_a_headers,
        "shipment",
        shipment_csv("SHP-TENANT-ROLLBACK"),
    )
    assert upload_response.status_code == 200
    job_id = upload_response.json()["ingestion_job_id"]

    with SessionLocal() as db:
        role = db.scalar(select(Role).where(Role.name == LOGISTICS_USER))
        assert role is not None
        tenant_b = Tenant(name="Tenant B", slug="tenant-b")
        user_b = User(
            email="rollback-b@test.local",
            full_name="Tenant B User",
            password_hash=hash_password("TestOnlyCredential1!"),
            is_active=True,
        )
        db.add_all([tenant_b, user_b])
        db.flush()
        db.add(
            TenantMembership(
                tenant_id=tenant_b.id,
                user_id=user_b.id,
                role_id=role.id,
                is_active=True,
            )
        )
        db.commit()

    login_response = client.post(
        "/api/v1/auth/login",
        json={"email": "rollback-b@test.local", "password": "TestOnlyCredential1!"},
    )
    assert login_response.status_code == 200
    tenant_b_headers = {
        "Authorization": f"Bearer {login_response.json()['access_token']}",
        "X-Tenant-Slug": "tenant-b",
    }

    rollback_response = client.post(
        f"/api/v1/ingestion/jobs/{job_id}/rollback",
        headers=tenant_b_headers,
    )

    assert rollback_response.status_code == 404
    with SessionLocal() as db:
        assert (
            db.scalar(
                select(func.count())
                .select_from(Shipment)
                .where(Shipment.shipment_id == "SHP-TENANT-ROLLBACK")
            )
            == 1
        )


def test_ingestion_history_does_not_leak_between_tenants(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, SessionLocal = client_and_session
    tenant_a_headers = login(client)
    upload_response = upload_csv(client, tenant_a_headers, "shipment", shipment_csv("SHP-TENANT-A"))
    assert upload_response.status_code == 200

    with SessionLocal() as db:
        role = db.scalar(select(Role).where(Role.name == LOGISTICS_USER))
        assert role is not None
        tenant_b = Tenant(name="Tenant B", slug="tenant-b")
        user_b = User(
            email="tenant-b@test.local",
            full_name="Tenant B User",
            password_hash=hash_password("TestOnlyCredential1!"),
            is_active=True,
        )
        db.add_all([tenant_b, user_b])
        db.flush()
        db.add(
            TenantMembership(
                tenant_id=tenant_b.id,
                user_id=user_b.id,
                role_id=role.id,
                is_active=True,
            )
        )
        db.commit()

    login_response = client.post(
        "/api/v1/auth/login",
        json={"email": "tenant-b@test.local", "password": "TestOnlyCredential1!"},
    )
    assert login_response.status_code == 200
    tenant_b_headers = {
        "Authorization": f"Bearer {login_response.json()['access_token']}",
        "X-Tenant-Slug": "tenant-b",
    }

    history_response = client.get("/api/v1/ingestion/jobs", headers=tenant_b_headers)
    assert history_response.status_code == 200
    assert history_response.json() == []


def test_threshold_upload(client_and_session: tuple[TestClient, sessionmaker[Session]]) -> None:
    client, SessionLocal = client_and_session
    response = upload_csv(
        client,
        login(client),
        "threshold",
        "\n".join(
            [
                "plant_code,material_code,threshold_days,warning_days",
                "JAM,COKING_COAL,7,10",
            ]
        ),
    )

    assert response.status_code == 200
    assert response.json()["summary_counts"]["created"] == 1
    assert response.json()["operational_summary"]["plants_detected"] == ["JAM"]
    with SessionLocal() as db:
        assert db.scalar(select(func.count()).select_from(PlantMaterialThreshold)) == 1


def test_threshold_upload_rejects_warning_days_before_critical_threshold(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, SessionLocal = client_and_session
    response = upload_csv(
        client,
        login(client),
        "threshold",
        "\n".join(
            [
                "plant_code,material_code,threshold_days,warning_days",
                "JAM,COKING_COAL,7,3",
            ]
        ),
    )

    assert response.status_code == 400
    detail = response.json()["detail"]
    assert detail["validation_errors"][0]["field_errors"][0]["field"] == "warning_days"
    assert "cannot be earlier" in detail["validation_errors"][0]["field_errors"][0]["reason"]
    assert detail["operational_summary"]["rows_rejected"] == 1
    with SessionLocal() as db:
        assert db.scalar(select(func.count()).select_from(PlantMaterialThreshold)) == 0


def test_duplicate_shipment_reupload_is_idempotent(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, SessionLocal = client_and_session
    headers = login(client)

    first_response = upload_csv(client, headers, "shipment", shipment_csv("SHP-REPEAT"))
    second_response = upload_csv(client, headers, "shipment", shipment_csv("SHP-REPEAT"))

    assert first_response.status_code == 200
    assert second_response.status_code == 200
    assert second_response.json()["summary_counts"]["unchanged"] == 1
    with SessionLocal() as db:
        assert db.scalar(select(func.count()).select_from(Shipment)) == 1


def test_shipment_upload_accepts_dispatched_state_and_date_only_eta(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, SessionLocal = client_and_session
    response = upload_csv(
        client,
        login(client),
        "shipment",
        "\n".join(
            [
                "shipment_id,plant_code,material_code,supplier_name,quantity_mt,"
                "planned_eta,current_eta,current_state,latest_update_at",
                "SHP-DISPATCHED,JAM,COKING_COAL,Supplier A,74000,"
                "2026-04-20,2026-04-21,dispatched,2026-04-15 09:00:00",
            ]
        ),
    )

    assert response.status_code == 200
    with SessionLocal() as db:
        shipment = db.scalar(select(Shipment).where(Shipment.shipment_id == "SHP-DISPATCHED"))
        assert shipment is not None
        assert shipment.current_state == "in_transit"


def test_shipment_upload_can_derive_current_eta_from_delay_days(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, SessionLocal = client_and_session
    response = upload_csv(
        client,
        login(client),
        "shipment",
        "\n".join(
            [
                "shipment_id,plant_code,material_code,supplier_name,quantity_mt,"
                "planned_eta,delay_days,current_state,latest_update_at",
                "SHP-DELAY-DAYS,JAM,COKING_COAL,Supplier A,74000,"
                "2026-04-20T08:00:00Z,2.5,in_transit,2026-04-15T09:00:00Z",
            ]
        ),
    )

    assert response.status_code == 200, response.json()
    with SessionLocal() as db:
        shipment = db.scalar(select(Shipment).where(Shipment.shipment_id == "SHP-DELAY-DAYS"))
        assert shipment is not None
        assert shipment.current_eta.isoformat() == "2026-04-22T20:00:00"


def test_operator_cannot_clear_uploaded_data(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, _ = client_and_session
    headers = login(client)
    response = client.delete("/api/v1/ingestion/uploads", headers=headers)
    assert response.status_code == 403


def test_admin_can_clear_uploaded_data(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, SessionLocal = client_and_session
    headers = admin_login(client)

    shipment_response = upload_csv(client, headers, "shipment", shipment_csv("SHP-CLEAR"))
    stock_response = upload_csv(
        client,
        headers,
        "stock",
        "\n".join(
            [
                "plant_code,material_code,on_hand_mt,quality_held_mt,available_to_consume_mt,daily_consumption_mt,snapshot_time",
                "JAM,COKING_COAL,10000,1000,9000,500,2026-04-15T08:00:00Z",
            ]
        ),
    )
    threshold_response = upload_csv(
        client,
        headers,
        "threshold",
        "\n".join(
            [
                "plant_code,material_code,threshold_days,warning_days",
                "JAM,COKING_COAL,7,10",
            ]
        ),
    )

    assert shipment_response.status_code == 200
    assert stock_response.status_code == 200
    assert threshold_response.status_code == 200

    delete_response = client.delete("/api/v1/ingestion/uploads", headers=headers)
    assert delete_response.status_code == 200
    body = delete_response.json()
    assert body["shipments"] == 1
    assert body["stock_snapshots"] == 1
    assert body["thresholds"] == 1
    assert body["ingestion_jobs"] == 3
    assert body["uploaded_files"] == 3
    assert body["operational_events"] == 2

    with SessionLocal() as db:
        assert db.scalar(select(func.count()).select_from(Shipment)) == 0
        assert db.scalar(select(func.count()).select_from(StockSnapshot)) == 0
        assert db.scalar(select(func.count()).select_from(PlantMaterialThreshold)) == 0
        assert db.scalar(select(func.count()).select_from(IngestionJob)) == 0
        assert db.scalar(select(func.count()).select_from(UploadedFile)) == 0
        assert db.scalar(select(func.count()).select_from(OperationalEvent)) == 0


def test_mapping_preview_and_manual_override_support_nonstandard_headers(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, SessionLocal = client_and_session
    headers = login(client)
    csv_body = "\n".join(
        [
            "OneDrive shipment export",
            "shipment reference,plant location,material name,vendor,qty,eta original,"
            "eta latest,status now,last updated",
            "SHP-MAP,JAM,COKING_COAL,Supplier A,74000,2026-04-20T08:00:00Z,"
            "2026-04-21T08:00:00Z,in_transit,2026-04-15T09:00:00Z",
        ]
    )

    preview_response = client.post(
        "/api/v1/ingestion/mapping-preview",
        headers=headers,
        data={"file_type": "shipment"},
        files={"file": ("mapping.csv", csv_body.encode(), "text/csv")},
    )
    assert preview_response.status_code == 200
    preview = preview_response.json()
    assert "shipment_id" in preview["required_fields"]
    assert preview["headers"][0] == "shipment reference"
    suggested = {item["source_header"]: item["suggested_field"] for item in preview["suggestions"]}
    assert suggested["shipment reference"] == "shipment_id"
    assert suggested["vendor"] == "supplier_name"

    upload_response = client.post(
        "/api/v1/ingestion/uploads",
        headers=headers,
        data={
            "file_type": "shipment",
            "mapping_overrides": json.dumps(
                {
                    "plant location": "plant_code",
                    "material name": "material_code",
                    "eta original": "planned_eta",
                    "eta latest": "current_eta",
                    "status now": "current_state",
                }
            ),
        },
        files={"file": ("mapping.csv", csv_body.encode(), "text/csv")},
    )
    assert upload_response.status_code == 200
    with SessionLocal() as db:
        shipment = db.scalar(select(Shipment))
        assert shipment is not None
        assert shipment.source_of_truth == "manual_upload"


def test_mapping_preview_targets_are_sheet_specific(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, _ = client_and_session
    headers = login(client)

    cases = {
        "stock": {
            "csv": "\n".join(
                [
                    "plant_code,material_code,on_hand_mt,quality_held_mt,"
                    "available_to_consume_mt,daily_consumption_mt,snapshot_time",
                    "JAM,COKING_COAL,100,5,95,10,2026-05-01",
                ]
            ),
            "expected": {"on_hand_mt", "quality_held_mt", "available_to_consume_mt"},
            "unexpected": {"planned_eta", "current_eta", "warning_days", "threshold_days"},
        },
        "threshold": {
            "csv": "\n".join(
                [
                    "plant_code,material_code,warning_days,threshold_days,"
                    "reserve_quantity_mt,quality_hold_quantity_mt",
                    "JAM,COKING_COAL,7,3,20,5",
                ]
            ),
            "expected": {
                "warning_days",
                "threshold_days",
                "reserve_quantity_mt",
                "quality_hold_quantity_mt",
            },
            "unexpected": {"vessel_name", "current_eta", "quality_held_mt"},
        },
        "shipment": {
            "csv": "\n".join(
                [
                    "shipment_id,plant_code,material_code,supplier_name,quantity_mt,"
                    "planned_eta,current_eta,last_eta,current_state,latest_update_at",
                    "SHP-1,JAM,COKING_COAL,Supplier A,10,2026-05-01,2026-05-02,"
                    "2026-04-30,in_transit,2026-04-29",
                ]
            ),
            "expected": {"planned_eta", "current_eta", "latest_eta", "origin_port"},
            "unexpected": {"warning_days", "threshold_days", "quality_held_mt"},
        },
    }

    for file_type, config in cases.items():
        response = client.post(
            "/api/v1/ingestion/mapping-preview",
            headers=headers,
            data={"file_type": file_type},
            files={"file": ("mapping.csv", config["csv"].encode(), "text/csv")},
        )

        assert response.status_code == 200, response.json()
        body = response.json()
        targets = set(body["required_fields"]) | set(body["optional_fields"])
        assert config["expected"].issubset(targets)
        assert targets.isdisjoint(config["unexpected"])


def test_threshold_preview_and_upload_recognize_reserve_and_quality_hold_quantities(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, SessionLocal = client_and_session
    headers = login(client)
    csv_body = "\n".join(
        [
            "plant_code,material_code,warning_days,critical_threshold_days,"
            "reserve_quantity_mt,quality_hold_quantity_mt,minimum_buffer_stock_mt,"
            "minimum_buffer_stock_days,stockout_alert_horizon_days",
            "JAM,COKING_COAL,7,3,500,150,300,2,10",
        ]
    )

    preview_response = client.post(
        "/api/v1/ingestion/mapping-preview",
        headers=headers,
        data={"file_type": "threshold"},
        files={"file": ("threshold.csv", csv_body.encode(), "text/csv")},
    )
    assert preview_response.status_code == 200, preview_response.json()
    suggestions = {
        item["source_header"]: (item["suggested_field"], item["confidence"])
        for item in preview_response.json()["suggestions"]
    }
    assert suggestions["reserve_quantity_mt"] == ("reserve_quantity_mt", "high")
    assert suggestions["quality_hold_quantity_mt"] == ("quality_hold_quantity_mt", "high")

    upload_response = upload_csv(client, headers, "threshold", csv_body)

    assert upload_response.status_code == 200, upload_response.json()
    with SessionLocal() as db:
        threshold = db.scalar(select(PlantMaterialThreshold))
        assert threshold is not None
        assert str(threshold.reserve_quantity_mt) == "500.00"
        assert str(threshold.quality_hold_quantity_mt) == "150.00"
        assert str(threshold.minimum_buffer_stock_mt) == "300.00"
        assert str(threshold.minimum_buffer_stock_days) == "2.00"
        assert str(threshold.stockout_alert_horizon_days) == "10.00"


def test_threshold_upload_without_new_quantity_fields_still_passes(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, SessionLocal = client_and_session
    response = upload_csv(
        client,
        login(client),
        "threshold",
        "\n".join(
            [
                "plant_code,material_code,warning_days,threshold_days",
                "JAM,COKING_COAL,7,3",
            ]
        ),
    )

    assert response.status_code == 200, response.json()
    with SessionLocal() as db:
        threshold = db.scalar(select(PlantMaterialThreshold))
        assert threshold is not None
        assert threshold.reserve_quantity_mt is None
        assert threshold.quality_hold_quantity_mt is None
        assert threshold.minimum_buffer_stock_mt is None


def test_inventory_quality_held_maps_separately_from_threshold_quality_hold(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, _ = client_and_session
    headers = login(client)
    csv_body = "\n".join(
        [
            "plant_code,material_code,on_hand_mt,quality_held_mt,available_to_consume_mt,daily_consumption_mt,snapshot_time",
            "JAM,COKING_COAL,100,5,95,10,2026-05-01",
        ]
    )

    response = client.post(
        "/api/v1/ingestion/mapping-preview",
        headers=headers,
        data={"file_type": "stock"},
        files={"file": ("stock.csv", csv_body.encode(), "text/csv")},
    )

    assert response.status_code == 200, response.json()
    suggestions = {
        item["source_header"]: (item["suggested_field"], item["confidence"])
        for item in response.json()["suggestions"]
    }
    assert suggestions["quality_held_mt"] == ("quality_held_mt", "high")
    assert "quality_hold_quantity_mt" not in response.json()["optional_fields"]


def test_shipment_current_and_last_eta_map_distinctly(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, SessionLocal = client_and_session
    headers = login(client)
    csv_body = "\n".join(
        [
            "shipment_id,plant_code,material_code,supplier_name,quantity_mt,"
            "origin,planned_eta,current_eta,last_eta,current_state,latest_update_at",
            "SHP-ETA-MAP,JAM,COKING_COAL,Supplier A,10,DEMO Destination A,2026-05-01,"
            "2026-05-03,2026-05-02,in_transit,2026-04-29",
        ]
    )

    preview_response = client.post(
        "/api/v1/ingestion/mapping-preview",
        headers=headers,
        data={"file_type": "shipment"},
        files={"file": ("shipment.csv", csv_body.encode(), "text/csv")},
    )
    assert preview_response.status_code == 200, preview_response.json()
    suggestions = {
        item["source_header"]: (item["suggested_field"], item["confidence"])
        for item in preview_response.json()["suggestions"]
    }
    assert suggestions["origin"] == ("origin_port", "high")
    assert suggestions["current_eta"] == ("current_eta", "high")
    assert suggestions["last_eta"] == ("latest_eta", "high")
    assert suggestions["planned_eta"] == ("planned_eta", "high")

    upload_response = upload_csv(client, headers, "shipment", csv_body)

    assert upload_response.status_code == 200, upload_response.json()
    with SessionLocal() as db:
        shipment = db.scalar(select(Shipment).where(Shipment.shipment_id == "SHP-ETA-MAP"))
        assert shipment is not None
        assert shipment.current_eta.isoformat() == "2026-05-03T00:00:00"
        assert shipment.latest_eta is not None
        assert shipment.latest_eta.isoformat() == "2026-05-02T00:00:00"


def test_founder_demo_csv_files_ingest_and_are_import_auditable(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, SessionLocal = client_and_session
    enable_demo_tenant(SessionLocal)
    headers = login(client)

    stock_response = upload_csv(
        client,
        headers,
        "stock",
        (FIXTURE_DIR / "demo_stock_snapshots.csv").read_text(),
    )
    shipment_response = upload_csv(
        client,
        headers,
        "shipment",
        (FIXTURE_DIR / "demo_inbound_shipments.csv").read_text(),
    )
    threshold_response = upload_csv(
        client,
        headers,
        "threshold",
        (FIXTURE_DIR / "demo_continuity_thresholds.csv").read_text(),
    )

    assert stock_response.status_code == 200, stock_response.json()
    assert shipment_response.status_code == 200, shipment_response.json()
    assert threshold_response.status_code == 200, threshold_response.json()
    assert stock_response.json()["rows_accepted"] == 4
    assert shipment_response.json()["rows_accepted"] == 6
    assert threshold_response.json()["rows_accepted"] == 4
    assert "DEMO-PLANT-A" in stock_response.json()["operational_summary"]["plants_detected"]
    assert (
        "DEMO-MATERIAL-A" in shipment_response.json()["operational_summary"]["materials_detected"]
    )
    assert (
        "DEMO-MAT-A-PROTECTIVE"
        in shipment_response.json()["operational_summary"]["shipments_detected"]
    )
    shipment_summary = shipment_response.json()["operational_summary"]
    assert any("created new master data" in warning for warning in shipment_summary["warnings"])
    assert shipment_summary["supplier_references_total"] >= 1
    assert shipment_summary["supplier_references_linked"] == shipment_summary[
        "supplier_references_total"
    ]
    assert shipment_summary["supplier_references_unlinked"] == 0

    detail_response = client.get(
        f"/api/v1/ingestion/jobs/{shipment_response.json()['ingestion_job_id']}",
        headers=headers,
    )

    assert detail_response.status_code == 200, detail_response.json()
    detail = detail_response.json()
    assert detail["created_records"] == 6
    assert all(record["rollback_safe"] for record in detail["record_references"])
    with SessionLocal() as db:
        assert (
            db.scalar(
                select(func.count())
                .select_from(Shipment)
                .where(Shipment.shipment_id.like("DEMO-%"))
            )
            == 6
        )
        assert db.scalar(select(func.count()).select_from(Supplier)) >= 1
        assert (
            db.scalar(
                select(func.count())
                .select_from(StockSnapshot)
                .join(Plant, Plant.id == StockSnapshot.plant_id)
                .where(Plant.code.like("DEMO_%"))
            )
            == 4
        )


def test_demo_coking_coal_upload_story_proves_time_phased_cover_and_history(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, SessionLocal = client_and_session
    enable_demo_tenant(SessionLocal)
    with SessionLocal() as db:
        tenant = db.scalar(select(Tenant).where(Tenant.slug == "tenant-a"))
        assert tenant is not None
        db.add(
            Supplier(
                tenant_id=tenant.id,
                code="DEMO-SUPPLIER-A",
                name="DEMO-SUPPLIER-A",
                primary_port="DEMO-ORIGIN-A",
                is_active=True,
            )
        )
        db.commit()

    headers = login(client)
    stock_response = upload_csv(
        client,
        headers,
        "stock",
        (FIXTURE_DIR / "demo_stock_snapshots.csv").read_text(),
    )
    shipment_response = upload_csv(
        client,
        headers,
        "shipment",
        (FIXTURE_DIR / "demo_inbound_shipments.csv").read_text(),
    )
    threshold_response = upload_csv(
        client,
        headers,
        "threshold",
        (FIXTURE_DIR / "demo_continuity_thresholds.csv").read_text(),
    )

    assert stock_response.status_code == 200, stock_response.json()
    assert shipment_response.status_code == 200, shipment_response.json()
    assert threshold_response.status_code == 200, threshold_response.json()
    operational_summary = shipment_response.json()["operational_summary"]
    assert operational_summary["supplier_references_total"] >= 2
    assert operational_summary["supplier_references_linked"] == operational_summary[
        "supplier_references_total"
    ]
    assert operational_summary["supplier_references_unlinked"] == 0
    assert operational_summary["onboarding_completeness_score"] == 100
    assert operational_summary["supplier_reliability_impact"] is None
    assert not any(
        "not linked to an existing supplier record" in warning
        for warning in operational_summary["warnings"]
    )

    with SessionLocal() as db:
        tenant = db.scalar(select(Tenant).where(Tenant.slug == "tenant-a"))
        assert tenant is not None
        plant = db.scalar(
            select(Plant).where(Plant.tenant_id == tenant.id, Plant.code == "DEMO_PLANT_A")
        )
        material = db.scalar(
            select(Material).where(
                Material.tenant_id == tenant.id,
                Material.code == "DEMO_MATERIAL_A",
            )
        )
        assert plant is not None
        assert material is not None
        threshold = db.scalar(
            select(PlantMaterialThreshold).where(
                PlantMaterialThreshold.tenant_id == tenant.id,
                PlantMaterialThreshold.plant_id == plant.id,
                PlantMaterialThreshold.material_id == material.id,
            )
        )
        assert threshold is not None
        threshold.minimum_buffer_stock_days = Decimal("2")
        threshold.minimum_buffer_stock_mt = Decimal("20")
        db.add(
            LineStopIncident(
                tenant_id=tenant.id,
                plant_id=plant.id,
                material_id=material.id,
                stopped_at=datetime(2026, 6, 5, 8, tzinfo=UTC),
                duration_hours=Decimal("8"),
                notes="Demo coking-coal historical line stop.",
            )
        )
        db.commit()

        context = RequestContext(
            tenant_id=tenant.id,
            tenant_slug=tenant.slug,
            role="tenant_admin",
            user_id=1,
        )
        detail = calculate_stock_cover_detail(db, context, plant.id, material.id)
        assert detail is not None
        cover = detail.time_phased_cover
        assert cover is not None
        assert cover.warning_date is not None
        assert cover.reserve_breach_date is not None
        assert cover.critical_breach_date is not None
        assert cover.interruption_date is not None
        assert cover.current_projected_warning_date is not None
        assert cover.current_projected_reserve_breach_date is not None
        assert cover.current_projected_critical_breach_date is not None
        assert cover.current_projected_interruption_date is not None
        statuses = {item.shipment_id: item.protection_status for item in cover.shipment_evaluations}
        assert statuses["DEMO-MAT-A-PROTECTIVE"] == "PROTECTIVE"
        assert statuses["DEMO-MAT-A-LATE"] == "LATE_AFTER_RESERVE"
        assert statuses["DEMO-MAT-A-UNLINKED"] == "TOO_LATE"
        assert "One or more inbound shipments are missing supplier master linkage." not in (
            cover.assumptions_used
        )

        report = build_historical_validation_report(db, context)
        coking_result = next(
            result
            for result in report.results
            if result.plant_id == plant.id and result.material_id == material.id
        )
        assert coking_result.predicted_warning_date is not None
        assert coking_result.lead_time_gained_hours is not None
        assert coking_result.lead_time_gained_hours > 0


def test_founder_demo_import_rollback_only_removes_demo_job_records(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, SessionLocal = client_and_session
    enable_demo_tenant(SessionLocal)
    headers = login(client)
    stock_response = upload_csv(
        client,
        headers,
        "stock",
        (FIXTURE_DIR / "demo_stock_snapshots.csv").read_text(),
    )
    shipment_response = upload_csv(
        client,
        headers,
        "shipment",
        (FIXTURE_DIR / "demo_inbound_shipments.csv").read_text(),
    )
    assert stock_response.status_code == 200
    assert shipment_response.status_code == 200

    rollback_response = client.post(
        f"/api/v1/ingestion/jobs/{shipment_response.json()['ingestion_job_id']}/rollback",
        headers=headers,
    )

    assert rollback_response.status_code == 200, rollback_response.json()
    assert rollback_response.json()["records_deleted"] == 6
    with SessionLocal() as db:
        assert (
            db.scalar(
                select(func.count())
                .select_from(Shipment)
                .where(Shipment.shipment_id.like("DEMO-%"))
            )
            == 0
        )
        assert db.scalar(select(func.count()).select_from(StockSnapshot)) == 4


def test_demo_prefixed_upload_rejected_for_non_demo_tenant(
    client_and_session: tuple[TestClient, sessionmaker[Session]],
) -> None:
    client, _ = client_and_session
    headers = login(client)

    response = upload_csv(
        client,
        headers,
        "shipment",
        (FIXTURE_DIR / "demo_inbound_shipments.csv").read_text(),
    )

    assert response.status_code == 400, response.json()
    body = response.json()["detail"]
    assert body["rows_accepted"] == 0
    assert body["rows_rejected"] == 6
    assert body["validation_errors"]
    first_error = body["validation_errors"][0]["field_errors"][0]
    assert "DEMO-" in first_error["reason"]
    assert "demo-enabled tenant" in first_error["suggested_fix"]


def shipment_csv(shipment_id: str) -> str:
    header = (
        "shipment_id,plant_code,material_code,supplier_name,quantity_mt,"
        "planned_eta,current_eta,current_state,latest_update_at"
    )
    row = (
        f"{shipment_id},JAM,COKING_COAL,Supplier A,74000,"
        "2026-04-20T08:00:00Z,2026-04-21T08:00:00Z,"
        "in_transit,2026-04-15T09:00:00Z"
    )
    return "\n".join(
        [
            header,
            row,
        ]
    )
