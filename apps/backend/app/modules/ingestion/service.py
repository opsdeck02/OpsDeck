from __future__ import annotations

import csv
import hashlib
import io
import json
import logging
import re
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import UTC, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from fastapi import HTTPException, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy import delete, func, select
from sqlalchemy.orm import Session

from app.models import (
    ExceptionCase,
    ExceptionComment,
    ImportJobRecord,
    IngestionJob,
    Material,
    OperationalEvent,
    Plant,
    PlantMaterialThreshold,
    Shipment,
    ShipmentUpdate,
    StockSnapshot,
    Supplier,
    Tenant,
    UploadedFile,
)
from app.models.enums import OperationalEventType, ShipmentState
from app.modules.exceptions.service import create_audit_log
from app.modules.ingestion.schemas import (
    FieldValidationError,
    HeaderMappingSuggestion,
    ImportJobDetailOut,
    ImportRecordReference,
    IngestionSummary,
    MappingPreviewOut,
    OperationalUnderstandingSummary,
    RejectionSummary,
    RollbackSummary,
    RowValidationError,
    SheetUploadResult,
    UploadResult,
    WorkbookPreviewOut,
    WorkbookSheetPreview,
    WorkbookUploadResult,
)
from app.modules.operational_events.service import (
    emit_inventory_stock_updated,
    emit_shipment_update_event,
)
from app.modules.suppliers.service import find_supplier_by_name
from app.modules.tenants.demo import is_demo_tenant
from app.schemas.context import RequestContext

logger = logging.getLogger(__name__)

SUPPORTED_FILE_TYPES = {"shipment", "stock", "threshold", "consumption"}
WORKBOOK_FILE_TYPES = {"stock", "shipment", "threshold", "consumption"}
WORKBOOK_TYPE_LABELS = {
    "stock": "Inventory",
    "shipment": "Inbound continuity",
    "threshold": "Thresholds",
    "consumption": "Consumption",
}
SHEET_TYPE_KEYWORDS = {
    "stock": {"stock", "inventory", "inventry", "currentstock", "stores"},
    "shipment": {"eta", "shipment", "shipments", "inbound", "dispatch", "tracking"},
    "threshold": {"threshold", "thresholds", "criticalcover", "warningcover"},
    "consumption": {"consumption", "usage", "consume", "consumed", "dailyusage"},
}
UPLOAD_DIR = Path("uploaded_files")

ALIASES = {
    "shipment_id": {"shipmentid", "shipment", "shipmentref", "reference", "shipmentreference"},
    "plant_code": {"plantcode", "plant", "plantid", "plantname"},
    "material_code": {"materialcode", "material", "materialid", "materialname"},
    "material_name": {"materialname"},
    "supplier_name": {"suppliername", "supplier", "vendor"},
    "quantity_mt": {
        "quantitymt",
        "quantity",
        "qtymt",
        "mt",
        "inboundqtytons",
        "inboundquantitytons",
    },
    "planned_eta": {"plannedeta", "originaleta", "eta", "dispatchdate", "shipmentdate"},
    "current_eta": {"currenteta", "latesteta", "revisedeta", "expectedarrivaldate", "arrivaldate"},
    "latest_eta": {"lasteta", "previouseta", "prioreta", "etaold", "etalast"},
    "delay_days": {"delaydays", "delay", "delays", "etadelaydays", "etadelay", "delayindays"},
    "current_state": {"currentstate", "state", "status", "shipmentstate", "shipmentstatus"},
    "source_of_truth": {"sourceoftruth", "source", "datasource"},
    "latest_update_at": {
        "latestupdateat",
        "lastupdatedat",
        "updatedat",
        "lastupdated",
        "eventtime",
    },
    "vessel_name": {"vesselname", "vessel", "shipname"},
    "imo_number": {"imonumber", "imo"},
    "mmsi": {"mmsi"},
    "origin_port": {"originport", "origin", "loadport"},
    "destination_port": {"destinationport", "destination", "dischargeport"},
    "eta_confidence": {"etaconfidence", "confidence"},
    "po_number": {"ponumber", "po", "purchaseorder", "purchaseordernumber"},
    "remarks": {"remarks", "remark", "comments", "notes"},
    "on_hand_mt": {"onhandmt", "onhand", "stockmt", "stock", "currentstocktons", "currentstock"},
    "quality_held_mt": {
        "qualityheldmt",
        "qualityheld",
        "heldmt",
        "blockedstocktons",
        "blockedstock",
    },
    "available_to_consume_mt": {
        "availabletoconsumemt",
        "availablemt",
        "available",
        "availableunrestrictedtons",
    },
    "daily_consumption_mt": {
        "dailyconsumptionmt",
        "dailyconsumption",
        "consumptionmt",
        "dailyconsumptiontons",
        "usage",
        "dailyusage",
        "consumption",
    },
    "snapshot_time": {"snapshottime", "snapshotat", "asof", "asoftime", "lastupdatedat"},
    "in_transit_open_tons": {"intransitopentons"},
    "days_to_line_stop": {"daystolinestop"},
    "risk_status": {"riskstatus"},
    "next_inbound_eta_days": {"nextinboundetadays"},
    "threshold_days": {
        "thresholddays",
        "threshold",
        "criticaldays",
        "criticalcoverdays",
        "criticalthresholddays",
    },
    "minimum_buffer_stock_mt": {"minimumbufferstockmt"},
    "reserve_quantity_mt": {"reservequantitymt", "reservequantity", "reserveqtymt", "reserveqty"},
    "quality_hold_quantity_mt": {
        "qualityholdquantitymt",
        "qualityholdquantity",
        "configuredqualityholdmt",
    },
    "minimum_buffer_stock_days": {"minimumbufferstockdays", "bufferstockdays", "reservedays"},
    "stockout_alert_horizon_days": {"stockoutalerthorizondays", "stockoutalertdays"},
    "warning_days": {"warningdays", "warning", "warningcoverdays", "mincoverdays"},
}

FIELD_LABELS = {
    "shipment_id": "Inbound reference",
    "plant_code": "Plant code/name",
    "material_code": "Material code/name",
    "supplier_name": "Reliability source",
    "quantity_mt": "Quantity MT",
    "planned_eta": "Planned ETA",
    "current_eta": "Current ETA",
    "latest_eta": "Previous ETA",
    "delay_days": "Delay days",
    "current_state": "Inbound continuity state",
    "source_of_truth": "Signal source",
    "latest_update_at": "Latest update time",
    "on_hand_mt": "On-hand MT",
    "quality_held_mt": "Quality-held MT",
    "available_to_consume_mt": "Available to consume MT",
    "daily_consumption_mt": "Daily consumption MT",
    "snapshot_time": "Snapshot time",
    "threshold_days": "Critical threshold days",
    "warning_days": "Warning days",
    "eta_confidence": "ETA confidence",
    "origin_port": "Origin port",
    "destination_port": "Destination port",
    "vessel_name": "Vessel name",
    "imo_number": "IMO number",
    "mmsi": "MMSI",
    "po_number": "PO number",
    "remarks": "Remarks",
    "minimum_buffer_stock_mt": "Minimum buffer stock MT",
    "reserve_quantity_mt": "Reserve quantity MT",
    "quality_hold_quantity_mt": "Quality hold quantity MT",
    "minimum_buffer_stock_days": "Minimum buffer stock days",
    "stockout_alert_horizon_days": "Stockout alert horizon days",
}

REQUIRED_FIELDS = {
    "shipment": {
        "shipment_id",
        "plant_code",
        "material_code",
        "supplier_name",
        "quantity_mt",
        "current_state",
        "latest_update_at",
    },
    "stock": {
        "plant_code",
        "material_code",
        "on_hand_mt",
        "quality_held_mt",
        "available_to_consume_mt",
        "daily_consumption_mt",
        "snapshot_time",
    },
    "threshold": {"plant_code", "material_code", "threshold_days", "warning_days"},
    "consumption": {"plant_code", "material_code", "daily_consumption_mt", "snapshot_time"},
}

HEADER_FIELDS_BY_FILE_TYPE = {
    "shipment": {
        "shipment_id",
        "plant_code",
        "material_code",
        "supplier_name",
        "quantity_mt",
        "planned_eta",
        "current_eta",
        "latest_eta",
        "delay_days",
        "current_state",
        "source_of_truth",
        "latest_update_at",
        "vessel_name",
        "imo_number",
        "mmsi",
        "origin_port",
        "destination_port",
        "eta_confidence",
    },
    "stock": {
        "plant_code",
        "material_code",
        "material_name",
        "on_hand_mt",
        "quality_held_mt",
        "available_to_consume_mt",
        "daily_consumption_mt",
        "snapshot_time",
        "in_transit_open_tons",
        "days_to_line_stop",
        "risk_status",
        "next_inbound_eta_days",
    },
    "threshold": {
        "plant_code",
        "material_code",
        "threshold_days",
        "warning_days",
        "minimum_buffer_stock_mt",
        "reserve_quantity_mt",
        "quality_hold_quantity_mt",
        "minimum_buffer_stock_days",
        "stockout_alert_horizon_days",
    },
    "consumption": {
        "plant_code",
        "material_code",
        "daily_consumption_mt",
        "snapshot_time",
    },
}


@dataclass
class ParsedRow:
    row_number: int
    data: dict[str, str]


@dataclass
class HeaderMatch:
    field: str | None
    confidence: str
    score: float
    alternatives: list[str]


class FieldValueError(ValueError):
    def __init__(self, field: str, reason: str, suggested_fix: str | None = None) -> None:
        self.field = field
        self.reason = reason
        self.suggested_fix = suggested_fix
        super().__init__(reason)


@dataclass
class ImportRecordChange:
    action: str
    record_type: str
    record_id: str
    record_reference: str | None
    rollback_safe: bool
    previous_value: dict[str, Any] | None = None
    new_value: dict[str, Any] | None = None


def process_upload(
    db: Session,
    context: RequestContext,
    current_user_id: int | None,
    file_type: str,
    upload: UploadFile,
    mapping_overrides: dict[str, str] | None = None,
) -> UploadResult:
    content = upload.file.read()
    return process_upload_content(
        db=db,
        context=context,
        current_user_id=current_user_id,
        file_type=file_type,
        filename=upload.filename or "upload",
        content=content,
        content_type=upload.content_type,
        mapping_overrides=mapping_overrides,
        source_of_truth="manual_upload",
    )


def process_upload_content(
    db: Session,
    context: RequestContext,
    current_user_id: int | None,
    file_type: str,
    filename: str,
    content: bytes,
    content_type: str | None = None,
    mapping_overrides: dict[str, str] | None = None,
    source_of_truth: str | None = None,
    event_source_id: int | None = None,
) -> UploadResult:
    file_type = file_type.lower().strip()
    if file_type not in SUPPORTED_FILE_TYPES:
        raise HTTPException(status_code=400, detail="Unsupported file type")

    if not content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    checksum = hashlib.sha256(content).hexdigest()
    uploaded_file = UploadedFile(
        tenant_id=context.tenant_id,
        original_filename=filename,
        storage_uri=save_upload(content, filename, checksum),
        content_type=content_type,
        file_size_bytes=len(content),
        checksum_sha256=checksum,
        uploaded_by_user_id=current_user_id,
        status="uploaded",
    )
    db.add(uploaded_file)
    db.flush()
    create_audit_log(
        db,
        _audit_context(context, current_user_id),
        action="ingestion.uploaded",
        entity_type="uploaded_file",
        entity_id=str(uploaded_file.id),
        metadata={
            "file_type": file_type,
            "filename": uploaded_file.original_filename,
            "checksum_sha256": checksum,
        },
    )

    job = IngestionJob(
        tenant_id=context.tenant_id,
        uploaded_file_id=uploaded_file.id,
        source_type=file_type,
        status="running",
        stage="parsing_file",
        started_at=datetime.now(UTC),
        records_total=0,
        records_succeeded=0,
        records_failed=0,
        metadata_json={
            "file_name": filename,
            "content_type": content_type,
            "mapping_overrides": mapping_overrides or {},
            "source_of_truth": source_of_truth,
        },
    )
    db.add(job)
    db.flush()

    try:
        job.stage = "parsing_file"
        rows = parse_upload(
            file_type,
            filename,
            content,
            mapping_overrides=mapping_overrides,
            source_of_truth=source_of_truth,
        )
        if not rows:
            raise ValueError("No data rows found in uploaded file")

        job.stage = "validating_rows"
        result = normalize_rows(
            db,
            context,
            file_type,
            rows,
            event_source_id=event_source_id or job.id,
            import_job_id=job.id,
        )
        result = result.model_copy(
            update={"upload_id": uploaded_file.id, "ingestion_job_id": job.id}
        )
        job.status = "completed" if result.rows_rejected == 0 else "completed_with_errors"
        job.stage = "completed"
        if result.rows_accepted == 0:
            job.status = "failed"
            job.stage = "failed"
        job.records_total = result.rows_received
        job.records_succeeded = result.rows_accepted
        job.records_failed = result.rows_rejected
        if result.validation_errors:
            job.error_message = json.dumps(
                [error.model_dump() for error in result.validation_errors]
            )
        job.metadata_json = {
            **(job.metadata_json or {}),
            "operational_summary": result.operational_summary.model_dump(),
            "blocking_errors": result.blocking_errors,
            "warnings": result.operational_summary.warnings,
        }
        uploaded_file.status = "failed" if result.rows_accepted == 0 else "processed"
        create_audit_log(
            db,
            _audit_context(context, current_user_id),
            action="ingestion.processed",
            entity_type="ingestion_job",
            entity_id=str(job.id),
            metadata={
                "file_type": file_type,
                "rows_received": result.rows_received,
                "rows_accepted": result.rows_accepted,
                "rows_rejected": result.rows_rejected,
            },
        )
        job.completed_at = datetime.now(UTC)
        db.commit()

        if result.rows_accepted == 0:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=result.model_dump())

        return result
    except HTTPException:
        raise
    except Exception as exc:
        db.rollback()
        # Persist failed metadata in a fresh transaction.
        db.add(uploaded_file)
        db.add(job)
        uploaded_file.status = "failed"
        job.status = "failed"
        job.stage = "failed"
        job.error_message = str(exc)
        job.completed_at = datetime.now(UTC)
        db.commit()
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc


def process_workbook_upload(
    db: Session,
    context: RequestContext,
    current_user_id: int | None,
    filename: str,
    content: bytes,
    content_type: str | None = None,
    sheet_configs: list[dict[str, Any]] | None = None,
    source_of_truth: str | None = "manual_upload",
) -> WorkbookUploadResult:
    if not content:
        raise HTTPException(status_code=400, detail="Uploaded workbook is empty")
    if Path(filename).suffix.lower() not in {".xlsx", ".xlsm"}:
        raise HTTPException(
            status_code=400, detail="Operational workbook upload requires XLSX/XLSM"
        )

    checksum = hashlib.sha256(content).hexdigest()
    uploaded_file = UploadedFile(
        tenant_id=context.tenant_id,
        original_filename=filename,
        storage_uri=save_upload(content, filename, checksum),
        content_type=content_type,
        file_size_bytes=len(content),
        checksum_sha256=checksum,
        uploaded_by_user_id=current_user_id,
        status="uploaded",
    )
    db.add(uploaded_file)
    db.flush()

    job = IngestionJob(
        tenant_id=context.tenant_id,
        uploaded_file_id=uploaded_file.id,
        source_type="workbook",
        status="running",
        stage="parsing_workbook",
        started_at=datetime.now(UTC),
        records_total=0,
        records_succeeded=0,
        records_failed=0,
        metadata_json={
            "file_name": filename,
            "content_type": content_type,
            "sheet_configs": sheet_configs or [],
            "source_of_truth": source_of_truth,
        },
    )
    db.add(job)
    db.flush()

    create_audit_log(
        db,
        _audit_context(context, current_user_id),
        action="ingestion.workbook_uploaded",
        entity_type="uploaded_file",
        entity_id=str(uploaded_file.id),
        metadata={"filename": filename, "checksum_sha256": checksum},
    )

    try:
        job.stage = "parsing_workbook"
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        config_by_sheet = workbook_config_by_sheet(sheet_configs or [])
        sheet_results: list[SheetUploadResult] = []
        ignored_sheets: list[str] = []
        all_validation_errors: list[RowValidationError] = []
        total_summary = IngestionSummary()

        for sheet in workbook.worksheets:
            rows = sheet_rows(sheet)
            if not rows:
                ignored_sheets.append(sheet.title)
                continue
            config = config_by_sheet.get(sheet.title)
            if not config or config["file_type"] == "ignore":
                ignored_sheets.append(sheet.title)
                continue

            file_type = config["file_type"]
            mapping_overrides = config.get("mapping_overrides") or {}
            try:
                job.stage = f"validating_{sheet.title}"
                parsed_rows = parse_xlsx_sheet_rows(
                    file_type,
                    rows,
                    mapping_overrides=mapping_overrides,
                    source_of_truth=source_of_truth,
                )
                if parsed_rows:
                    result = normalize_rows(
                        db,
                        context,
                        file_type,
                        parsed_rows,
                        event_source_id=job.id,
                        import_job_id=job.id,
                    )
                else:
                    result = empty_upload_result(file_type)
                db.flush()
            except ValueError as exc:
                result = blocking_sheet_result(sheet.title, file_type, str(exc))

            sheet_results.append(
                SheetUploadResult(
                    sheet_name=sheet.title,
                    file_type=file_type,
                    status=sheet_status(result),
                    rows_received=result.rows_received,
                    rows_accepted=result.rows_accepted,
                    rows_rejected=result.rows_rejected,
                    validation_errors=result.validation_errors,
                    top_rejection_reasons=result.top_rejection_reasons,
                    blocking_errors=result.blocking_errors,
                    summary_counts=result.summary_counts,
                    operational_summary=result.operational_summary,
                )
            )
            all_validation_errors.extend(prefix_sheet_errors(sheet.title, result.validation_errors))
            total_summary.created += result.summary_counts.created
            total_summary.updated += result.summary_counts.updated
            total_summary.unchanged += result.summary_counts.unchanged

        rows_received = sum(sheet.rows_received for sheet in sheet_results)
        rows_accepted = sum(sheet.rows_accepted for sheet in sheet_results)
        rows_rejected = sum(sheet.rows_rejected for sheet in sheet_results)
        blocking_errors = [
            f"{sheet.sheet_name}: {error}"
            for sheet in sheet_results
            for error in sheet.blocking_errors
        ]
        top_rejections = top_rejection_reasons(all_validation_errors)
        result = WorkbookUploadResult(
            upload_id=uploaded_file.id,
            ingestion_job_id=job.id,
            rows_received=rows_received,
            rows_accepted=rows_accepted,
            rows_rejected=rows_rejected,
            validation_errors=all_validation_errors,
            top_rejection_reasons=top_rejections,
            blocking_errors=blocking_errors,
            summary_counts=total_summary,
            operational_summary=aggregate_workbook_operational_summary(
                sheet_results,
                ignored_sheets,
            ),
            sheet_results=sheet_results,
            ignored_sheets=ignored_sheets,
        )

        job.status = (
            "completed" if rows_rejected == 0 and not blocking_errors else "completed_with_errors"
        )
        job.stage = "completed"
        if rows_accepted == 0:
            job.status = "failed"
            job.stage = "failed"
        job.records_total = rows_received
        job.records_succeeded = rows_accepted
        job.records_failed = rows_rejected
        if all_validation_errors or blocking_errors:
            job.error_message = json.dumps(
                {
                    "blocking_errors": blocking_errors,
                    "validation_errors": [error.model_dump() for error in all_validation_errors],
                    "sheet_results": [sheet.model_dump() for sheet in sheet_results],
                }
            )
        job.metadata_json = {
            **(job.metadata_json or {}),
            "operational_summary": result.operational_summary.model_dump(),
            "blocking_errors": blocking_errors,
            "warnings": result.operational_summary.warnings,
            "ignored_sheets": ignored_sheets,
        }
        uploaded_file.status = "failed" if rows_accepted == 0 else "processed"
        job.completed_at = datetime.now(UTC)
        create_audit_log(
            db,
            _audit_context(context, current_user_id),
            action="ingestion.workbook_processed",
            entity_type="ingestion_job",
            entity_id=str(job.id),
            metadata={
                "rows_received": rows_received,
                "rows_accepted": rows_accepted,
                "rows_rejected": rows_rejected,
                "sheets": len(sheet_results),
                "ignored_sheets": ignored_sheets,
            },
        )
        db.commit()

        if rows_accepted == 0:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=result.model_dump())
        return result
    except HTTPException:
        raise
    except Exception as exc:
        db.rollback()
        db.add(uploaded_file)
        db.add(job)
        uploaded_file.status = "failed"
        job.status = "failed"
        job.stage = "failed"
        job.error_message = str(exc)
        job.completed_at = datetime.now(UTC)
        db.commit()
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc


def _audit_context(context: RequestContext, current_user_id: int | None) -> RequestContext:
    return context.model_copy(update={"user_id": current_user_id})


def save_upload(content: bytes, filename: str, checksum: str) -> str:
    UPLOAD_DIR.mkdir(exist_ok=True)
    safe_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", filename)
    path = UPLOAD_DIR / f"{checksum[:12]}_{safe_name}"
    path.write_bytes(content)
    return str(path)


def preview_workbook_mapping(filename: str, content: bytes) -> WorkbookPreviewOut:
    if Path(filename).suffix.lower() not in {".xlsx", ".xlsm"}:
        raise HTTPException(
            status_code=400, detail="Operational workbook preview requires XLSX/XLSM"
        )
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheets: list[WorkbookSheetPreview] = []
    ignored_empty_sheets: list[str] = []
    for sheet in workbook.worksheets:
        rows = sheet_rows(sheet)
        if not rows:
            ignored_empty_sheets.append(sheet.title)
            continue
        suggested_file_type = suggest_sheet_file_type(sheet.title, rows)
        previews = {
            file_type: build_mapping_preview_from_rows(file_type, rows)
            for file_type in sorted(WORKBOOK_FILE_TYPES)
        }
        row_count = detected_data_row_count(rows, suggested_file_type or "stock")
        sheets.append(
            WorkbookSheetPreview(
                sheet_name=sheet.title,
                hidden=sheet.sheet_state != "visible",
                row_count=row_count,
                suggested_file_type=suggested_file_type,
                suggested_label=WORKBOOK_TYPE_LABELS.get(suggested_file_type or ""),
                previews=previews,
            )
        )
    return WorkbookPreviewOut(
        file_name=filename,
        sheets=sheets,
        ignored_empty_sheets=ignored_empty_sheets,
    )


def preview_header_mapping(
    file_type: str,
    filename: str,
    content: bytes,
) -> MappingPreviewOut:
    file_type = file_type.lower().strip()
    if file_type not in SUPPORTED_FILE_TYPES:
        raise HTTPException(status_code=400, detail="Unsupported file type")
    headers = extract_headers(file_type, filename, content)
    return build_mapping_preview(file_type, headers)


def build_mapping_preview_from_rows(
    file_type: str, rows: list[tuple[Any, ...]]
) -> MappingPreviewOut:
    header_index = detect_header_row(file_type, rows)
    headers = ["" if value is None else str(value) for value in rows[header_index]]
    return build_mapping_preview(file_type, headers)


def build_mapping_preview(file_type: str, headers: list[str]) -> MappingPreviewOut:
    suggestions = [build_header_mapping_suggestion(file_type, header) for header in headers]
    required_fields = sorted(REQUIRED_FIELDS[file_type])
    optional_fields = sorted(
        field
        for field in HEADER_FIELDS_BY_FILE_TYPE[file_type]
        if field not in REQUIRED_FIELDS[file_type]
    )
    mapped_required_fields = sorted(
        {
            suggestion.suggested_field
            for suggestion in suggestions
            if suggestion.suggested_field in REQUIRED_FIELDS[file_type]
        }
    )
    if file_type == "shipment":
        mapped_required_fields = sorted({*mapped_required_fields, "source_of_truth"})
    missing_required_fields = sorted(REQUIRED_FIELDS[file_type] - set(mapped_required_fields))
    blocking_errors = [
        f"{field_label(field)} is required but no uploaded column is mapped."
        for field in missing_required_fields
    ]
    return MappingPreviewOut(
        file_type=file_type,
        headers=headers,
        required_fields=required_fields,
        optional_fields=optional_fields,
        suggestions=suggestions,
        mapped_required_fields=mapped_required_fields,
        missing_required_fields=missing_required_fields,
        blocking_errors=blocking_errors,
    )


def delete_uploaded_data(db: Session, tenant_id: int) -> dict[str, int]:
    uploaded_files = list(
        db.scalars(select(UploadedFile).where(UploadedFile.tenant_id == tenant_id))
    )
    deleted_counts = {
        "uploaded_files": len(uploaded_files),
        "ingestion_jobs": int(
            db.scalar(
                select(func.count(IngestionJob.id)).where(IngestionJob.tenant_id == tenant_id)
            )
            or 0
        ),
        "shipments": int(
            db.scalar(select(func.count(Shipment.id)).where(Shipment.tenant_id == tenant_id)) or 0
        ),
        "stock_snapshots": int(
            db.scalar(
                select(func.count(StockSnapshot.id)).where(StockSnapshot.tenant_id == tenant_id)
            )
            or 0
        ),
        "thresholds": int(
            db.scalar(
                select(func.count(PlantMaterialThreshold.id)).where(
                    PlantMaterialThreshold.tenant_id == tenant_id
                )
            )
            or 0
        ),
        "shipment_updates": int(
            db.scalar(
                select(func.count(ShipmentUpdate.id)).where(ShipmentUpdate.tenant_id == tenant_id)
            )
            or 0
        ),
        "exceptions": int(
            db.scalar(
                select(func.count(ExceptionCase.id)).where(ExceptionCase.tenant_id == tenant_id)
            )
            or 0
        ),
        "operational_events": int(
            db.scalar(
                select(func.count(OperationalEvent.id)).where(
                    OperationalEvent.tenant_id == tenant_id
                )
            )
            or 0
        ),
    }

    for uploaded_file in uploaded_files:
        if uploaded_file.storage_uri:
            try:
                Path(uploaded_file.storage_uri).unlink(missing_ok=True)
            except OSError:
                pass

    db.execute(delete(ExceptionComment).where(ExceptionComment.tenant_id == tenant_id))
    db.execute(delete(ExceptionCase).where(ExceptionCase.tenant_id == tenant_id))
    db.execute(delete(OperationalEvent).where(OperationalEvent.tenant_id == tenant_id))
    db.execute(delete(ShipmentUpdate).where(ShipmentUpdate.tenant_id == tenant_id))
    db.execute(delete(Shipment).where(Shipment.tenant_id == tenant_id))
    db.execute(delete(StockSnapshot).where(StockSnapshot.tenant_id == tenant_id))
    db.execute(delete(PlantMaterialThreshold).where(PlantMaterialThreshold.tenant_id == tenant_id))
    db.execute(delete(IngestionJob).where(IngestionJob.tenant_id == tenant_id))
    db.execute(delete(UploadedFile).where(UploadedFile.tenant_id == tenant_id))
    db.commit()
    return deleted_counts


def get_import_job_detail(
    db: Session,
    context: RequestContext,
    import_job_id: int,
) -> ImportJobDetailOut:
    job = get_import_job_or_404(db, context.tenant_id, import_job_id)
    uploaded_file = db.get(UploadedFile, job.uploaded_file_id) if job.uploaded_file_id else None
    records = import_job_records(db, context.tenant_id, job.id)
    metadata = job.metadata_json or {}
    row_errors = row_errors_from_job(job)
    summary = operational_summary_from_metadata(metadata)
    return ImportJobDetailOut(
        import_job_id=job.id,
        upload_id=job.uploaded_file_id,
        file_name=uploaded_file.original_filename if uploaded_file else None,
        import_type=job.source_type,
        status=job.status,
        stage=job.stage,
        total_rows=job.records_total,
        accepted_rows=job.records_succeeded,
        rejected_rows=job.records_failed,
        created_records=sum(1 for record in records if record.action == "created"),
        updated_records=sum(1 for record in records if record.action == "updated"),
        unchanged_records=sum(1 for record in records if record.action == "unchanged"),
        warnings=list(metadata.get("warnings") or []),
        row_level_errors=row_errors,
        operational_summary=summary,
        record_references=[
            ImportRecordReference(
                record_type=record.record_type,
                record_id=record.record_id,
                record_reference=record.record_reference,
                action=record.action,
                rollback_safe=record.rollback_safe,
                rollback_status=record.rollback_status,
            )
            for record in records
        ],
        started_at=job.started_at.isoformat() if job.started_at else None,
        completed_at=job.completed_at.isoformat() if job.completed_at else None,
        uploaded_at=uploaded_file.created_at.isoformat() if uploaded_file else None,
        source_metadata={
            key: value
            for key, value in metadata.items()
            if key
            not in {
                "operational_summary",
                "warnings",
                "blocking_errors",
            }
        },
    )


def rollback_import_job(
    db: Session,
    context: RequestContext,
    import_job_id: int,
) -> RollbackSummary:
    job = get_import_job_or_404(db, context.tenant_id, import_job_id)
    if job.status == "rolled_back":
        return RollbackSummary(
            import_job_id=job.id,
            rollback_status="already_rolled_back",
            warnings=["This import job was already rolled back."],
        )

    records = import_job_records(db, context.tenant_id, job.id)
    deleted = 0
    skipped = 0
    preserved = 0
    skipped_reasons: list[str] = []
    warnings: list[str] = []

    for record in records:
        if record.action != "created" or not record.rollback_safe:
            preserved += 1
            record.rollback_status = "preserved"
            skipped_reasons.append(
                f"{record.record_type} {record.record_reference or record.record_id} was "
                "not created by this import or cannot be safely reverted."
            )
            continue
        if was_record_touched_after_import(db, context.tenant_id, record):
            skipped += 1
            record.rollback_status = "skipped_later_import"
            skipped_reasons.append(
                f"{record.record_type} {record.record_reference or record.record_id} was "
                "preserved because another import also touched it."
            )
            continue
        if delete_import_created_record(db, context.tenant_id, job.id, record):
            deleted += 1
            record.rollback_status = "deleted"
        else:
            skipped += 1
            record.rollback_status = "skipped_missing"
            skipped_reasons.append(
                f"{record.record_type} {record.record_reference or record.record_id} was "
                "not found during rollback."
            )

    if preserved:
        warnings.append(
            "Updated pre-existing records were preserved. Exact update rollback is not "
            "available; re-upload corrected data if needed."
        )
    job.status = "rolled_back"
    job.stage = "rolled_back"
    job.metadata_json = {
        **(job.metadata_json or {}),
        "rollback_summary": {
            "records_deleted": deleted,
            "records_skipped": skipped,
            "records_preserved": preserved,
            "skipped_reasons": skipped_reasons,
            "warnings": warnings,
        },
    }
    db.commit()
    return RollbackSummary(
        import_job_id=job.id,
        rollback_status="rolled_back",
        records_deleted=deleted,
        records_skipped=skipped,
        records_preserved=preserved,
        skipped_reasons=skipped_reasons,
        warnings=warnings,
    )


def reprocess_import_job(
    db: Session,
    context: RequestContext,
    current_user_id: int | None,
    import_job_id: int,
) -> UploadResult | WorkbookUploadResult:
    job = get_import_job_or_404(db, context.tenant_id, import_job_id)
    uploaded_file = db.get(UploadedFile, job.uploaded_file_id) if job.uploaded_file_id else None
    if uploaded_file is None or not uploaded_file.storage_uri:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                "This import job cannot be reprocessed because its uploaded file is unavailable."
            ),
        )
    path = Path(uploaded_file.storage_uri)
    if not path.exists():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="This import job cannot be reprocessed because its stored file is missing.",
        )
    metadata = job.metadata_json or {}
    content = path.read_bytes()
    if job.source_type == "workbook":
        sheet_configs = metadata.get("sheet_configs")
        if not isinstance(sheet_configs, list):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Workbook reprocess requires the original sheet configuration.",
            )
        return process_workbook_upload(
            db=db,
            context=context,
            current_user_id=current_user_id,
            filename=uploaded_file.original_filename,
            content=content,
            content_type=uploaded_file.content_type,
            sheet_configs=sheet_configs,
            source_of_truth=metadata.get("source_of_truth") or "manual_upload",
        )
    return process_upload_content(
        db=db,
        context=context,
        current_user_id=current_user_id,
        file_type=job.source_type,
        filename=uploaded_file.original_filename,
        content=content,
        content_type=uploaded_file.content_type,
        mapping_overrides=metadata.get("mapping_overrides") or {},
        source_of_truth=metadata.get("source_of_truth") or "manual_upload",
    )


def get_import_job_or_404(db: Session, tenant_id: int, import_job_id: int) -> IngestionJob:
    job = db.get(IngestionJob, import_job_id)
    if job is None or job.tenant_id != tenant_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Import job not found")
    return job


def import_job_records(
    db: Session,
    tenant_id: int,
    import_job_id: int,
) -> list[ImportJobRecord]:
    return list(
        db.scalars(
            select(ImportJobRecord)
            .where(
                ImportJobRecord.tenant_id == tenant_id,
                ImportJobRecord.import_job_id == import_job_id,
            )
            .order_by(ImportJobRecord.id)
        )
    )


def row_errors_from_job(job: IngestionJob) -> list[RowValidationError]:
    if not job.error_message:
        return []
    try:
        parsed = json.loads(job.error_message)
    except json.JSONDecodeError:
        return [
            RowValidationError(
                row_number=0,
                errors=[job.error_message],
                field_errors=[],
            )
        ]
    if isinstance(parsed, dict):
        parsed = parsed.get("validation_errors") or []
    if not isinstance(parsed, list):
        return []
    return [RowValidationError(**item) for item in parsed if isinstance(item, dict)]


def operational_summary_from_metadata(
    metadata: dict[str, Any],
) -> OperationalUnderstandingSummary | None:
    summary = metadata.get("operational_summary")
    if not isinstance(summary, dict):
        return None
    return OperationalUnderstandingSummary(**summary)


def was_record_touched_after_import(
    db: Session,
    tenant_id: int,
    record: ImportJobRecord,
) -> bool:
    later_record = db.scalar(
        select(ImportJobRecord)
        .where(
            ImportJobRecord.tenant_id == tenant_id,
            ImportJobRecord.record_type == record.record_type,
            ImportJobRecord.record_id == record.record_id,
            ImportJobRecord.id > record.id,
        )
        .limit(1)
    )
    return later_record is not None


def delete_import_created_record(
    db: Session,
    tenant_id: int,
    import_job_id: int,
    record: ImportJobRecord,
) -> bool:
    if record.record_type == "shipment":
        shipment = db.get(Shipment, int(record.record_id))
        if shipment is None or shipment.tenant_id != tenant_id:
            return False
        db.execute(
            delete(OperationalEvent).where(
                OperationalEvent.tenant_id == tenant_id,
                OperationalEvent.source_id == import_job_id,
                OperationalEvent.shipment_id == shipment.id,
            )
        )
        db.delete(shipment)
        return True
    if record.record_type == "stock_snapshot":
        snapshot = db.get(StockSnapshot, int(record.record_id))
        if snapshot is None or snapshot.tenant_id != tenant_id:
            return False
        db.execute(
            delete(OperationalEvent).where(
                OperationalEvent.tenant_id == tenant_id,
                OperationalEvent.source_id == import_job_id,
                OperationalEvent.plant_id == snapshot.plant_id,
                OperationalEvent.material_id == snapshot.material_id,
            )
        )
        db.delete(snapshot)
        return True
    if record.record_type == "threshold":
        threshold = db.get(PlantMaterialThreshold, int(record.record_id))
        if threshold is None or threshold.tenant_id != tenant_id:
            return False
        db.delete(threshold)
        return True
    return False


def parse_upload(
    file_type: str,
    filename: str,
    content: bytes,
    mapping_overrides: dict[str, str] | None = None,
    source_of_truth: str | None = None,
) -> list[ParsedRow]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        return parse_csv(
            file_type,
            content,
            mapping_overrides=mapping_overrides,
            source_of_truth=source_of_truth,
        )
    if suffix in {".xlsx", ".xlsm"}:
        return parse_xlsx(
            file_type,
            content,
            mapping_overrides=mapping_overrides,
            source_of_truth=source_of_truth,
        )
    raise ValueError("Only CSV, XLSX, and XLSM uploads are supported")


def extract_headers(file_type: str, filename: str, content: bytes) -> list[str]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        text = content.decode("utf-8-sig")
        rows = list(csv.reader(io.StringIO(text)))
        header_index = detect_header_row(file_type, rows)
        headers = rows[header_index] if rows else []
        if not headers:
            raise ValueError("Missing header row")
        return ["" if value is None else str(value) for value in headers]
    if suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise ValueError("Missing header row")
        header_index = detect_header_row(file_type, rows)
        return ["" if value is None else str(value) for value in rows[header_index]]
    raise ValueError("Only CSV, XLSX, and XLSM uploads are supported")


def parse_csv(
    file_type: str,
    content: bytes,
    mapping_overrides: dict[str, str] | None = None,
    source_of_truth: str | None = None,
) -> list[ParsedRow]:
    text = content.decode("utf-8-sig")
    rows = list(csv.reader(io.StringIO(text)))
    if not rows:
        raise ValueError("Missing header row")
    header_index = detect_header_row(file_type, rows)
    headers = ["" if value is None else str(value) for value in rows[header_index]]
    log_missing_required_headers(file_type, headers, header_index)
    data_rows = (dict(zip(headers, row, strict=False)) for row in rows[header_index + 1 :])
    return normalize_records(
        file_type,
        headers,
        data_rows,
        start_row=header_index + 2,
        mapping_overrides=mapping_overrides,
        source_of_truth=source_of_truth,
    )


def parse_xlsx(
    file_type: str,
    content: bytes,
    mapping_overrides: dict[str, str] | None = None,
    source_of_truth: str | None = None,
) -> list[ParsedRow]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Missing header row")
    header_index = detect_header_row(file_type, rows)
    headers = ["" if value is None else str(value) for value in rows[header_index]]
    log_missing_required_headers(file_type, headers, header_index)
    records = (dict(zip(headers, row, strict=False)) for row in rows[header_index + 1 :])
    return normalize_records(
        file_type,
        headers,
        records,
        start_row=header_index + 2,
        mapping_overrides=mapping_overrides,
        source_of_truth=source_of_truth,
    )


def parse_xlsx_sheet_rows(
    file_type: str,
    rows: list[tuple[Any, ...]],
    mapping_overrides: dict[str, str] | None = None,
    source_of_truth: str | None = None,
) -> list[ParsedRow]:
    header_index = detect_header_row(file_type, rows)
    headers = ["" if value is None else str(value) for value in rows[header_index]]
    log_missing_required_headers(file_type, headers, header_index)
    records = (dict(zip(headers, row, strict=False)) for row in rows[header_index + 1 :])
    return normalize_records(
        file_type,
        headers,
        records,
        start_row=header_index + 2,
        mapping_overrides=mapping_overrides,
        source_of_truth=source_of_truth,
    )


def sheet_rows(sheet: Any) -> list[tuple[Any, ...]]:
    rows: list[tuple[Any, ...]] = []
    for row in sheet.iter_rows(values_only=True):
        if any(cell is not None and str(cell).strip() for cell in row):
            rows.append(tuple(row))
    return rows


def detected_data_row_count(rows: list[tuple[Any, ...]], file_type: str) -> int:
    if not rows:
        return 0
    header_index = detect_header_row(file_type, rows)
    return sum(
        1
        for row in rows[header_index + 1 :]
        if any(cell is not None and str(cell).strip() for cell in row)
    )


def suggest_sheet_file_type(sheet_name: str, rows: list[tuple[Any, ...]]) -> str | None:
    normalized_name = normalize_header_token(sheet_name)
    for file_type, keywords in SHEET_TYPE_KEYWORDS.items():
        if any(keyword in normalized_name for keyword in keywords):
            return file_type

    scores = {
        file_type: header_row_score(file_type, rows[detect_header_row(file_type, rows)])
        for file_type in WORKBOOK_FILE_TYPES
    }
    best_type, best_score = max(scores.items(), key=lambda item: item[1])
    if best_score >= 4:
        return best_type
    return None


def detect_header_row(file_type: str, rows: list[Iterable[Any]]) -> int:
    best_index = 0
    best_score = -1.0
    for index, row in enumerate(rows[:10]):
        headers = ["" if value is None else str(value) for value in row]
        score = header_row_score(file_type, headers)
        if score > best_score:
            best_score = score
            best_index = index
    if best_score <= 0:
        return 0
    return best_index


def log_missing_required_headers(file_type: str, headers: Iterable[str], header_index: int) -> None:
    matched_fields = {
        match.field
        for header in headers
        if (
            match := best_header_match(str(header), required_fields=REQUIRED_FIELDS[file_type])
        ).field
    }
    missing = sorted(REQUIRED_FIELDS[file_type] - matched_fields)
    if missing:
        logger.warning(
            "Parser detected header row %s for %s but required columns are missing: %s",
            header_index + 1,
            file_type,
            ", ".join(missing),
        )


def header_row_score(file_type: str, headers: Iterable[str]) -> float:
    required = REQUIRED_FIELDS[file_type]
    matched_fields = {
        match.field
        for header in headers
        if (match := best_header_match(header, required_fields=required)).field
    }
    non_empty_headers = sum(1 for header in headers if str(header).strip())
    return len(matched_fields) * 2 + min(non_empty_headers, len(required)) * 0.1


def normalize_records(
    file_type: str,
    headers: Iterable[str],
    records: Iterable[dict[str, Any]],
    start_row: int,
    mapping_overrides: dict[str, str] | None = None,
    source_of_truth: str | None = None,
) -> list[ParsedRow]:
    overrides = {str(key): str(value) for key, value in (mapping_overrides or {}).items() if value}
    header_map = {
        header: overrides.get(header) or canonical_header(header, file_type=file_type)
        for header in headers
    }
    missing_required_mapping = missing_required_mapped_fields(
        file_type,
        header_map,
        source_of_truth=source_of_truth,
    )
    if missing_required_mapping:
        labels = ", ".join(field_label(field) for field in missing_required_mapping)
        raise ValueError(f"Missing required mapping: {labels}")
    parsed_rows: list[ParsedRow] = []
    for index, record in enumerate(records, start=start_row):
        normalized = {}
        for raw_header, value in record.items():
            canonical = header_map.get(raw_header)
            if canonical:
                normalized_value = "" if value is None else str(value).strip()
                if canonical in normalized and normalized[canonical]:
                    continue
                normalized[canonical] = normalized_value
        if not any(value for value in normalized.values()):
            continue
        if source_of_truth and not normalized.get("source_of_truth"):
            normalized["source_of_truth"] = source_of_truth
        parsed_rows.append(ParsedRow(row_number=index, data=normalized))
    return parsed_rows


def missing_required_mapped_fields(
    file_type: str,
    header_map: dict[str, str | None],
    source_of_truth: str | None = None,
) -> list[str]:
    mapped_fields = {field for field in header_map.values() if field}
    required_fields = set(REQUIRED_FIELDS[file_type])
    if source_of_truth:
        required_fields.discard("source_of_truth")
    return sorted(required_fields - mapped_fields)


def field_label(field: str) -> str:
    return FIELD_LABELS.get(field, field.replace("_", " "))


def canonical_header(header: str, file_type: str | None = None) -> str | None:
    required_fields = HEADER_FIELDS_BY_FILE_TYPE.get(file_type) if file_type else None
    return best_header_match(header, required_fields=required_fields).field


def build_header_mapping_suggestion(file_type: str, header: str) -> HeaderMappingSuggestion:
    suggestion = best_header_match(header, required_fields=HEADER_FIELDS_BY_FILE_TYPE[file_type])
    return HeaderMappingSuggestion(
        source_header=header,
        suggested_field=suggestion.field,
        confidence=suggestion.confidence,
        alternatives=suggestion.alternatives,
    )


def best_header_match(header: str, required_fields: set[str] | None = None) -> HeaderMatch:
    compact = normalize_header_token(header)
    candidates = (
        {field: aliases for field, aliases in ALIASES.items() if field in required_fields}
        if required_fields is not None
        else ALIASES
    )
    ranked: list[tuple[float, str]] = []
    for canonical, aliases in candidates.items():
        canonical_compact = canonical.replace("_", "")
        if compact == canonical_compact or compact in aliases:
            ranked.append((1.0, canonical))
            continue

        score = similarity_score(compact, canonical, aliases)
        if score > 0:
            ranked.append((score, canonical))

    ranked.sort(key=lambda item: item[0], reverse=True)
    alternatives = [field for _, field in ranked[1:4]]
    if not ranked:
        return HeaderMatch(field=None, confidence="low", score=0.0, alternatives=[])

    best_score, best_field = ranked[0]
    confidence = "low"
    if best_score >= 0.9:
        confidence = "high"
    elif best_score >= 0.6:
        confidence = "medium"

    if best_score < 0.35:
        return HeaderMatch(
            field=None, confidence="low", score=best_score, alternatives=alternatives
        )
    return HeaderMatch(
        field=best_field,
        confidence=confidence,
        score=best_score,
        alternatives=alternatives,
    )


def normalize_header_token(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def similarity_score(compact: str, canonical: str, aliases: set[str]) -> float:
    canonical_compact = canonical.replace("_", "")
    candidates = {canonical_compact, *aliases}
    best = 0.0
    for candidate in candidates:
        if compact in candidate or candidate in compact:
            best = max(best, 0.72)
        compact_tokens = split_header_tokens(compact)
        candidate_tokens = split_header_tokens(candidate)
        overlap = len(compact_tokens & candidate_tokens)
        if overlap:
            token_score = overlap / max(len(candidate_tokens), 1)
            best = max(best, 0.45 + token_score * 0.4)
    return best


def split_header_tokens(value: str) -> set[str]:
    parts = re.findall(r"[a-z]+|\d+", value.lower())
    return {part for part in parts if part}


def normalize_rows(
    db: Session,
    context: RequestContext,
    file_type: str,
    rows: list[ParsedRow],
    event_source_id: int | None = None,
    import_job_id: int | None = None,
) -> UploadResult:
    validation_errors: list[RowValidationError] = []
    summary = IngestionSummary()
    demo_upload_allowed = is_demo_tenant(db, context.tenant_id)

    for row in rows:
        if not demo_upload_allowed:
            demo_errors = demo_prefix_field_errors(row.data)
            if demo_errors:
                validation_errors.append(build_row_validation_error(row.row_number, demo_errors))
                continue

        field_errors = missing_required_field_errors(file_type, row.data)
        if field_errors:
            validation_errors.append(build_row_validation_error(row.row_number, field_errors))
            continue

        try:
            if file_type == "shipment":
                change = upsert_shipment(db, context, row.data, event_source_id=event_source_id)
            elif file_type == "stock":
                change = upsert_stock_snapshot(
                    db,
                    context,
                    row.data,
                    event_source_id=event_source_id,
                )
            elif file_type == "consumption":
                change = upsert_consumption(
                    db,
                    context,
                    row.data,
                    event_source_id=event_source_id,
                )
            else:
                change = upsert_threshold(db, context, row.data)
        except FieldValueError as exc:
            validation_errors.append(
                build_row_validation_error(
                    row.row_number,
                    [
                        FieldValidationError(
                            field=exc.field,
                            reason=exc.reason,
                            suggested_fix=exc.suggested_fix,
                        )
                    ],
                )
            )
            continue
        except ValueError as exc:
            validation_errors.append(
                build_row_validation_error(row.row_number, [generic_field_error(exc)])
            )
            continue

        if import_job_id is not None:
            record_import_change(db, context.tenant_id, import_job_id, change)
        setattr(summary, change.action, getattr(summary, change.action) + 1)

    rows_accepted = summary.created + summary.updated + summary.unchanged
    rows_rejected = len(validation_errors)
    return UploadResult(
        upload_id=0,
        ingestion_job_id=0,
        file_type=file_type,
        rows_received=len(rows),
        rows_accepted=rows_accepted,
        rows_rejected=rows_rejected,
        validation_errors=validation_errors,
        top_rejection_reasons=top_rejection_reasons(validation_errors),
        summary_counts=summary,
        operational_summary=build_operational_summary(
            file_type=file_type,
            rows=rows,
            rows_received=len(rows),
            rows_accepted=rows_accepted,
            rows_rejected=rows_rejected,
            summary=summary,
            warnings=upload_warnings(db, context, file_type, rows),
            supplier_validation=supplier_onboarding_validation(db, context, file_type, rows),
        ),
    )


def empty_upload_result(file_type: str) -> UploadResult:
    return UploadResult(
        upload_id=0,
        ingestion_job_id=0,
        file_type=file_type,
        rows_received=0,
        rows_accepted=0,
        rows_rejected=0,
        validation_errors=[],
        summary_counts=IngestionSummary(),
        operational_summary=OperationalUnderstandingSummary(
            file_received=True,
            next_recommended_action=(
                "No continuity rows were detected. Check the sheet and upload again."
            ),
        ),
    )


def blocking_sheet_result(sheet_name: str, file_type: str, error: str) -> UploadResult:
    return UploadResult(
        upload_id=0,
        ingestion_job_id=0,
        file_type=file_type,
        rows_received=0,
        rows_accepted=0,
        rows_rejected=1,
        validation_errors=[],
        blocking_errors=[error],
        top_rejection_reasons=[RejectionSummary(reason=error, count=1)],
        summary_counts=IngestionSummary(),
        operational_summary=OperationalUnderstandingSummary(
            file_received=True,
            rows_rejected=1,
            warnings=[f"{sheet_name}: {error}"],
            next_recommended_action=(
                "Fix the blocking mapping or sheet issue, then reprocess the workbook."
            ),
        ),
    )


def sheet_status(result: UploadResult) -> str:
    if result.blocking_errors or result.rows_accepted == 0 and result.rows_rejected > 0:
        return "failed"
    if result.rows_rejected > 0:
        return "completed_with_errors"
    return "completed"


def prefix_sheet_errors(
    sheet_name: str, errors: list[RowValidationError]
) -> list[RowValidationError]:
    prefixed: list[RowValidationError] = []
    for error in errors:
        field_errors = [
            FieldValidationError(
                field=field_error.field,
                reason=f"{sheet_name}: {field_error.reason}",
                suggested_fix=field_error.suggested_fix,
            )
            for field_error in error.field_errors
        ]
        prefixed.append(
            RowValidationError(
                row_number=error.row_number,
                errors=[field_error.reason for field_error in field_errors] or error.errors,
                field_errors=field_errors,
            )
        )
    return prefixed


def build_operational_summary(
    *,
    file_type: str,
    rows: list[ParsedRow],
    rows_received: int,
    rows_accepted: int,
    rows_rejected: int,
    summary: IngestionSummary,
    warnings: list[str] | None = None,
    supplier_validation: dict[str, object] | None = None,
) -> OperationalUnderstandingSummary:
    warnings = warnings or []
    supplier_validation = supplier_validation or {}
    return OperationalUnderstandingSummary(
        file_received=True,
        rows_detected=rows_received,
        rows_accepted=rows_accepted,
        rows_rejected=rows_rejected,
        records_created=summary.created,
        records_updated=summary.updated,
        records_unchanged=summary.unchanged,
        plants_detected=unique_row_values(rows, "plant_code"),
        materials_detected=unique_row_values(rows, "material_code"),
        shipments_detected=(
            unique_row_values(rows, "shipment_id") if file_type == "shipment" else []
        ),
        suppliers_detected=(
            unique_row_values(rows, "supplier_name") if file_type == "shipment" else []
        ),
        risks_or_exposures_generated=None,
        refreshed_operational_visibility=rows_accepted > 0,
        warnings=warnings,
        next_recommended_action=next_upload_action(rows_accepted, rows_rejected, warnings),
        supplier_references_total=int(supplier_validation.get("total", 0)),
        supplier_references_linked=int(supplier_validation.get("linked", 0)),
        supplier_references_unlinked=int(supplier_validation.get("unlinked", 0)),
        onboarding_completeness_score=int(supplier_validation.get("score", 100)),
        supplier_reliability_impact=(
            str(supplier_validation["impact"]) if supplier_validation.get("impact") else None
        ),
    )


def record_import_change(
    db: Session,
    tenant_id: int,
    import_job_id: int,
    change: ImportRecordChange,
) -> None:
    recorded_at = datetime.now(UTC)
    db.add(
        ImportJobRecord(
            tenant_id=tenant_id,
            import_job_id=import_job_id,
            record_type=change.record_type,
            record_id=change.record_id,
            record_reference=change.record_reference,
            action=change.action,
            rollback_safe=change.rollback_safe,
            previous_value=change.previous_value,
            new_value=change.new_value,
            created_at=recorded_at,
            updated_at=recorded_at,
        )
    )


def aggregate_workbook_operational_summary(
    sheet_results: list[SheetUploadResult],
    ignored_sheets: list[str],
) -> OperationalUnderstandingSummary:
    summaries = [sheet.operational_summary for sheet in sheet_results]
    rows_received = sum(summary.rows_detected for summary in summaries)
    rows_accepted = sum(summary.rows_accepted for summary in summaries)
    rows_rejected = sum(summary.rows_rejected for summary in summaries)
    warnings = [warning for summary in summaries for warning in summary.warnings]
    if ignored_sheets:
        warnings.append("Ignored workbook sheets: " + ", ".join(sorted(set(ignored_sheets))))
    return OperationalUnderstandingSummary(
        file_received=True,
        rows_detected=rows_received,
        rows_accepted=rows_accepted,
        rows_rejected=rows_rejected,
        records_created=sum(summary.records_created for summary in summaries),
        records_updated=sum(summary.records_updated for summary in summaries),
        records_unchanged=sum(summary.records_unchanged for summary in summaries),
        plants_detected=unique_values(
            value for summary in summaries for value in summary.plants_detected
        ),
        materials_detected=unique_values(
            value for summary in summaries for value in summary.materials_detected
        ),
        shipments_detected=unique_values(
            value for summary in summaries for value in summary.shipments_detected
        ),
        suppliers_detected=unique_values(
            value for summary in summaries for value in summary.suppliers_detected
        ),
        risks_or_exposures_generated=None,
        refreshed_operational_visibility=rows_accepted > 0,
        warnings=warnings,
        next_recommended_action=next_upload_action(rows_accepted, rows_rejected, warnings),
        supplier_references_total=sum(summary.supplier_references_total for summary in summaries),
        supplier_references_linked=sum(summary.supplier_references_linked for summary in summaries),
        supplier_references_unlinked=sum(
            summary.supplier_references_unlinked for summary in summaries
        ),
        onboarding_completeness_score=aggregate_onboarding_score(summaries),
        supplier_reliability_impact=aggregate_supplier_reliability_impact(summaries),
    )


def unique_row_values(rows: list[ParsedRow], field: str) -> list[str]:
    return unique_values(row.data.get(field) for row in rows)


def unique_values(values: Iterable[str | None]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        normalized = (value or "").strip()
        if not normalized:
            continue
        key = normalize_text_token(normalized)
        if key in seen:
            continue
        seen.add(key)
        result.append(normalized)
    return result[:12]


def upload_warnings(
    db: Session,
    context: RequestContext,
    file_type: str,
    rows: list[ParsedRow],
) -> list[str]:
    warnings: list[str] = []
    if file_type == "shipment":
        known_suppliers = {
            normalize_text_token(name)
            for name in db.scalars(
                select(Supplier.name)
                .where(Supplier.tenant_id == context.tenant_id)
                .where(Supplier.is_active.is_(True))
            )
            if name
        }
        unknown_supplier_names = [
            supplier
            for supplier in unique_row_values(rows, "supplier_name")
            if normalize_text_token(supplier) not in known_suppliers
        ]
        if unknown_supplier_names:
            warnings.append(
                "Reliability source was accepted from the upload but is not linked to "
                f"an existing supplier record: {', '.join(unknown_supplier_names[:5])}."
            )
            warnings.append(
                "Supplier reliability calculations will be uncalibrated for unlinked "
                "shipment suppliers until supplier master records are created and linked."
            )
    return warnings


def supplier_onboarding_validation(
    db: Session,
    context: RequestContext,
    file_type: str,
    rows: list[ParsedRow],
) -> dict[str, object]:
    if file_type != "shipment":
        return {"total": 0, "linked": 0, "unlinked": 0, "score": 100, "impact": None}
    supplier_names = unique_row_values(rows, "supplier_name")
    if not supplier_names:
        return {"total": 0, "linked": 0, "unlinked": 0, "score": 100, "impact": None}
    known_suppliers = {
        normalize_text_token(name)
        for name in db.scalars(
            select(Supplier.name)
            .where(Supplier.tenant_id == context.tenant_id)
            .where(Supplier.is_active.is_(True))
        )
        if name
    }
    linked = sum(
        1 for supplier in supplier_names if normalize_text_token(supplier) in known_suppliers
    )
    total = len(supplier_names)
    unlinked = total - linked
    score = int((linked / total) * 100) if total else 100
    impact = None
    if unlinked:
        impact = (
            f"{unlinked} of {total} uploaded supplier references are not linked to supplier "
            "master records; supplier reliability context will be unknown for those shipments."
        )
    return {
        "total": total,
        "linked": linked,
        "unlinked": unlinked,
        "score": score,
        "impact": impact,
    }


def aggregate_onboarding_score(summaries: list[OperationalUnderstandingSummary]) -> int:
    total = sum(summary.supplier_references_total for summary in summaries)
    if total <= 0:
        return 100
    linked = sum(summary.supplier_references_linked for summary in summaries)
    return int((linked / total) * 100)


def aggregate_supplier_reliability_impact(
    summaries: list[OperationalUnderstandingSummary],
) -> str | None:
    impacts = [
        summary.supplier_reliability_impact
        for summary in summaries
        if summary.supplier_reliability_impact
    ]
    if not impacts:
        return None
    return " ".join(impacts)


def next_upload_action(
    rows_accepted: int,
    rows_rejected: int,
    warnings: list[str],
) -> str:
    if rows_accepted == 0 and rows_rejected > 0:
        return "Fix rejected rows or required mappings, then upload again."
    if rows_rejected > 0:
        return (
            "Review rejected rows, correct the file, and re-upload if those records "
            "should contribute to continuity visibility."
        )
    if warnings:
        return (
            "Review upload warnings, then check the Risk Workspace once the "
            "continuity signal chain refreshes."
        )
    if rows_accepted > 0:
        return "Check the Risk Workspace to confirm the refreshed continuity signals."
    return "Upload a continuity signal file with operational rows."


def workbook_config_by_sheet(configs: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    parsed: dict[str, dict[str, Any]] = {}
    for config in configs:
        sheet_name = str(config.get("sheet_name") or "").strip()
        file_type = str(config.get("file_type") or "ignore").strip().lower()
        if not sheet_name:
            continue
        if file_type == "ignore":
            parsed[sheet_name] = {"file_type": "ignore", "mapping_overrides": {}}
            continue
        if file_type not in WORKBOOK_FILE_TYPES:
            raise ValueError(f"Unsupported workbook sheet type: {file_type}")
        mapping_overrides = config.get("mapping_overrides") or {}
        if not isinstance(mapping_overrides, dict):
            raise ValueError(f"Mapping overrides for {sheet_name} must be an object")
        parsed[sheet_name] = {
            "file_type": file_type,
            "mapping_overrides": {
                str(key): str(value) for key, value in mapping_overrides.items() if value
            },
        }
    return parsed


def missing_required_field_errors(
    file_type: str, data: dict[str, str]
) -> list[FieldValidationError]:
    errors = [
        FieldValidationError(
            field=field,
            reason=missing_required_reason(field),
            suggested_fix=missing_required_fix(field),
        )
        for field in sorted(REQUIRED_FIELDS[file_type])
        if not data.get(field)
    ]
    if file_type == "shipment" and not data.get("planned_eta") and not data.get("current_eta"):
        errors.append(
            FieldValidationError(
                field="current_eta",
                reason="Planned ETA or Current ETA is required for inbound continuity.",
                suggested_fix="Provide either Planned ETA or Current ETA for this row.",
            )
        )
    return errors


def missing_required_reason(field: str) -> str:
    if field == "plant_code":
        return (
            "Plant code/name is missing. OpsDeck cannot link this row to a "
            "plant-material continuity record."
        )
    if field == "material_code":
        return (
            "Material code/name is missing. OpsDeck cannot identify which material "
            "this operational signal belongs to."
        )
    if field == "shipment_id":
        return (
            "Inbound reference is missing. OpsDeck cannot safely deduplicate or update "
            "this inbound row."
        )
    if field == "supplier_name":
        return (
            "Reliability source is missing. OpsDeck cannot explain which external "
            "dependency produced this inbound signal."
        )
    if field in {"planned_eta", "current_eta", "latest_update_at", "snapshot_time"}:
        return (
            f"{field_label(field)} is missing. OpsDeck needs this timing to assess "
            "continuity freshness."
        )
    if field in {
        "quantity_mt",
        "on_hand_mt",
        "quality_held_mt",
        "available_to_consume_mt",
        "daily_consumption_mt",
    }:
        return (
            f"{field_label(field)} is missing. OpsDeck needs this quantity to calculate "
            "operating cover."
        )
    return f"{field_label(field)} is required but was blank or unmapped."


def missing_required_fix(field: str) -> str:
    return f"Map a column to {field_label(field)} and make sure every operational row has a value."


def build_row_validation_error(
    row_number: int,
    field_errors: list[FieldValidationError],
) -> RowValidationError:
    return RowValidationError(
        row_number=row_number,
        errors=[error.reason for error in field_errors],
        field_errors=field_errors,
    )


def demo_prefix_field_errors(data: dict[str, str]) -> list[FieldValidationError]:
    fields_to_check = (
        "plant_code",
        "material_code",
        "shipment_id",
        "supplier_name",
    )
    errors: list[FieldValidationError] = []
    for field in fields_to_check:
        value = data.get(field)
        if value and value.strip().upper().startswith("DEMO-"):
            errors.append(
                FieldValidationError(
                    field=field,
                    reason=(
                        f"{field_label(field)} uses a DEMO- reference reserved for "
                        "demo-enabled tenants."
                    ),
                    suggested_fix=(
                        "Use a demo-enabled tenant for pilot sample files, or remove "
                        "DEMO- references before uploading real customer data."
                    ),
                )
            )
    return errors


def generic_field_error(exc: ValueError) -> FieldValidationError:
    return FieldValidationError(
        field="row",
        reason=str(exc),
        suggested_fix="Check the row values and upload again.",
    )


def top_rejection_reasons(errors: list[RowValidationError]) -> list[RejectionSummary]:
    counts: dict[str, int] = {}
    for row_error in errors:
        for error in row_error.field_errors:
            counts[error.reason] = counts.get(error.reason, 0) + 1
        if not row_error.field_errors:
            for reason in row_error.errors:
                counts[reason] = counts.get(reason, 0) + 1
    ranked = sorted(counts.items(), key=lambda item: item[1], reverse=True)
    return [RejectionSummary(reason=reason, count=count) for reason, count in ranked[:5]]


def resolve_plant(db: Session, tenant_id: int, value: str) -> Plant:
    normalized_value = normalize_text_token(value)
    plants = list(db.scalars(select(Plant).where(Plant.tenant_id == tenant_id)))
    plant = next(
        (
            item
            for item in plants
            if normalize_text_token(item.code) == normalized_value
            or normalize_text_token(item.name) == normalized_value
        ),
        None,
    )
    if plant is None:
        requested_index = infer_plant_index(value)
        if requested_index is not None:
            plant = next(
                (
                    item
                    for item in plants
                    if infer_plant_index(item.code) == requested_index
                    or infer_plant_index(item.name) == requested_index
                ),
                None,
            )
            if plant is not None and should_adopt_uploaded_plant_name(
                plant, requested_index, value
            ):
                plant.name = value.strip()
    if plant is None:
        tenant = db.get(Tenant, tenant_id)
        plant_count = int(
            db.scalar(select(func.count(Plant.id)).where(Plant.tenant_id == tenant_id)) or 0
        )
        if tenant and tenant.max_plants is not None and plant_count >= tenant.max_plants:
            raise ValueError(f"Tenant is limited to {tenant.max_plants} plants")
        plant = Plant(
            tenant_id=tenant_id,
            code=build_plant_code(value),
            name=value.strip(),
            location=None,
        )
        db.add(plant)
        db.flush()
    return plant


def resolve_material(db: Session, tenant_id: int, value: str) -> Material:
    normalized_value = normalize_text_token(value)
    materials = list(db.scalars(select(Material).where(Material.tenant_id == tenant_id)))
    material = next(
        (
            item
            for item in materials
            if normalize_text_token(item.code) == normalized_value
            or normalize_text_token(item.name) == normalized_value
        ),
        None,
    )
    if material is not None:
        return material

    code = build_material_code(value)
    material = db.scalar(
        select(Material).where(Material.tenant_id == tenant_id, Material.code == code)
    )
    if material is not None:
        return material

    material = Material(
        tenant_id=tenant_id,
        code=code,
        name=value.strip(),
        category=infer_material_category(value),
        uom="MT",
    )
    db.add(material)
    db.flush()
    return material


def parse_decimal(value: str, field: str, *, positive: bool = False) -> Decimal:
    raw_value = "" if value is None else str(value).strip()
    normalized = re.sub(
        r"\s*(mt|mts|metric\s*tons?|tonnes?|tons?|t)\s*$",
        "",
        raw_value,
        flags=re.IGNORECASE,
    ).strip()
    try:
        parsed = Decimal(normalized.replace(",", ""))
    except (InvalidOperation, AttributeError) as exc:
        raise FieldValueError(
            field,
            f"{field_label(field)} could not be interpreted from uploaded value.",
            f"Use a number for {field_label(field)}, for example 12500 or 12,500 MT.",
        ) from exc
    if positive and parsed <= 0:
        raise FieldValueError(
            field,
            f"{field_label(field)} must be greater than zero.",
            f"Enter a positive number for {field_label(field)}.",
        )
    return parsed


def parse_optional_decimal(
    data: dict[str, str],
    field: str,
    *,
    fallback_field: str | None = None,
) -> Decimal | None:
    if data.get(field):
        return parse_decimal(data[field], field)
    if fallback_field and data.get(fallback_field):
        return parse_decimal(data[fallback_field], fallback_field)
    return None


def parse_datetime(value: str, field: str) -> datetime:
    raw_value = "" if value is None else str(value).strip()
    if not raw_value:
        raise FieldValueError(
            field,
            f"{field_label(field)} is blank.",
            f"Add a date/time value for {field_label(field)}.",
        )
    parsed: datetime | None = None
    try:
        normalized = raw_value.replace("Z", "+00:00")
        parsed = datetime.fromisoformat(normalized)
    except ValueError:
        parsed = parse_common_datetime(raw_value)
    if parsed is None:
        raise FieldValueError(
            field,
            f"{field_label(field)} could not be interpreted from uploaded value.",
            "Use a recognizable date such as 2026-05-14, 14/05/2026, or 14-05-2026 09:30.",
        )
    if parsed.tzinfo is None:
        return parsed.replace(tzinfo=UTC)
    return parsed


def parse_common_datetime(value: str) -> datetime | None:
    formats = [
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
        "%d-%m-%Y %H:%M:%S",
        "%d-%m-%Y %H:%M",
        "%d-%m-%Y",
        "%d.%m.%Y %H:%M:%S",
        "%d.%m.%Y %H:%M",
        "%d.%m.%Y",
        "%d/%m/%y %H:%M",
        "%d/%m/%y",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y/%m/%d",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M",
        "%m/%d/%Y",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(value, fmt).replace(tzinfo=UTC)
        except ValueError:
            continue
    return None


def resolve_current_eta(data: dict[str, str], planned_eta: datetime) -> datetime:
    if data.get("current_eta"):
        return parse_datetime(data["current_eta"], "current_eta")
    if not data.get("delay_days"):
        return planned_eta
    delay_days = parse_decimal(data["delay_days"], "delay_days")
    return planned_eta + timedelta(days=float(delay_days))


def parse_state(value: str) -> ShipmentState:
    normalized = value.strip().lower().replace(" ", "_").replace("-", "_")
    state_aliases = {
        "dispatched": ShipmentState.IN_TRANSIT,
        "dispatch": ShipmentState.IN_TRANSIT,
        "intransit": ShipmentState.IN_TRANSIT,
        "in_transit": ShipmentState.IN_TRANSIT,
        "arrived": ShipmentState.AT_PORT,
        "received": ShipmentState.DELIVERED,
        "complete": ShipmentState.DELIVERED,
        "completed": ShipmentState.DELIVERED,
    }
    if normalized in state_aliases:
        return state_aliases[normalized]
    try:
        return ShipmentState(normalized)
    except ValueError as exc:
        raise FieldValueError(
            "current_state",
            "Inbound continuity state is not recognized.",
            "Use a known state such as in_transit, at_port, delivered, or delayed.",
        ) from exc


def values_equal(left: Any, right: Any) -> bool:
    if isinstance(left, datetime) and isinstance(right, datetime):
        return normalize_datetime(left) == normalize_datetime(right)
    return left == right


def normalize_datetime(value: datetime) -> datetime:
    if value.tzinfo is None:
        return value.replace(tzinfo=UTC)
    return value.astimezone(UTC)


def json_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return normalize_datetime(value).isoformat()
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, ShipmentState):
        return value.value
    return value


def shipment_event_value(attrs: dict[str, Any]) -> dict[str, Any]:
    return {
        "current_eta": json_value(attrs.get("current_eta")),
        "current_state": json_value(attrs.get("current_state")),
        "planned_eta": json_value(attrs.get("planned_eta")),
        "latest_update_at": json_value(attrs.get("latest_update_at")),
        "quantity_mt": json_value(attrs.get("quantity_mt")),
        "supplier_name": attrs.get("supplier_name"),
        "source_of_truth": attrs.get("source_of_truth"),
    }


def stock_event_value(attrs: dict[str, Any], snapshot_time: datetime) -> dict[str, Any]:
    return {
        "snapshot_time": json_value(snapshot_time),
        "on_hand_mt": json_value(attrs.get("on_hand_mt")),
        "quality_held_mt": json_value(attrs.get("quality_held_mt")),
        "available_to_consume_mt": json_value(attrs.get("available_to_consume_mt")),
        "daily_consumption_mt": json_value(attrs.get("daily_consumption_mt")),
    }


def normalize_text_token(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def infer_plant_index(value: str) -> int | None:
    normalized = value.strip().lower()
    number_match = re.search(r"([0-9]+)\s*$", normalized)
    if number_match:
        return int(number_match.group(1))

    letter_match = re.search(r"([a-z])\s*$", normalized)
    if letter_match:
        return ord(letter_match.group(1)) - ord("a") + 1

    return None


def build_plant_code(value: str) -> str:
    compact = re.sub(r"[^A-Za-z0-9]+", "_", value.strip().upper()).strip("_")
    return compact[:40] or "PLANT"


def build_material_code(value: str) -> str:
    compact = re.sub(r"[^A-Za-z0-9]+", "_", value.strip().upper()).strip("_")
    return compact[:40] or "MATERIAL"


def infer_material_category(value: str) -> str:
    normalized = value.strip().lower()
    if "coal" in normalized:
        return "coal"
    if "ore" in normalized:
        return "ore"
    if "lime" in normalized or "flux" in normalized:
        return "flux"
    return "raw"


def should_adopt_uploaded_plant_name(
    plant: Plant, requested_index: int, uploaded_value: str
) -> bool:
    normalized_uploaded = uploaded_value.strip()
    if not normalized_uploaded:
        return False
    return plant.name.strip().lower() == f"plant {requested_index}".lower()


def upsert_shipment(
    db: Session,
    context: RequestContext,
    data: dict[str, str],
    event_source_id: int | None = None,
) -> ImportRecordChange:
    plant = resolve_plant(db, context.tenant_id, data["plant_code"])
    material = resolve_material(db, context.tenant_id, data["material_code"])
    state = parse_state(data["current_state"])
    planned_eta = parse_datetime(
        data.get("planned_eta") or data["current_eta"],
        "planned_eta" if data.get("planned_eta") else "current_eta",
    )
    current_eta = resolve_current_eta(data, planned_eta)
    previous_eta = (
        parse_datetime(data["latest_eta"], "latest_eta") if data.get("latest_eta") else None
    )
    latest_update_at = parse_datetime(data["latest_update_at"], "latest_update_at")
    eta_confidence = (
        parse_decimal(data["eta_confidence"], "eta_confidence")
        if data.get("eta_confidence")
        else None
    )
    supplier = find_supplier_by_name(db, context.tenant_id, data["supplier_name"])

    shipment = db.scalar(
        select(Shipment).where(
            Shipment.tenant_id == context.tenant_id,
            Shipment.shipment_id == data["shipment_id"],
        )
    )
    attrs = {
        "material_id": material.id,
        "plant_id": plant.id,
        "supplier_id": supplier.id if supplier else None,
        "supplier_name": data["supplier_name"],
        "quantity_mt": parse_decimal(data["quantity_mt"], "quantity_mt", positive=True),
        "vessel_name": data.get("vessel_name") or None,
        "imo_number": data.get("imo_number") or None,
        "mmsi": data.get("mmsi") or None,
        "origin_port": data.get("origin_port") or None,
        "destination_port": data.get("destination_port") or None,
        "planned_eta": planned_eta,
        "current_eta": current_eta,
        "latest_eta": previous_eta,
        "delay_days": int(parse_decimal(data["delay_days"], "delay_days"))
        if data.get("delay_days")
        else None,
        "eta_confidence": eta_confidence,
        "current_state": state,
        "source_of_truth": data["source_of_truth"],
        "latest_update_at": latest_update_at,
    }

    new_value = shipment_event_value(attrs)
    source_reference = data.get("source_of_truth")
    if shipment is None:
        shipment = Shipment(tenant_id=context.tenant_id, shipment_id=data["shipment_id"], **attrs)
        db.add(shipment)
        db.flush()
        emit_shipment_update_event(
            db,
            tenant_id=context.tenant_id,
            event_type=OperationalEventType.SHIPMENT_MILESTONE_UPDATED,
            occurred_at=latest_update_at,
            source_reference=source_reference,
            shipment_id=shipment.id,
            shipment_reference=shipment.shipment_id,
            plant_id=plant.id,
            plant_reference=plant.code,
            material_id=material.id,
            material_reference=material.code,
            supplier_id=shipment.supplier_id,
            supplier_reference=shipment.supplier_name,
            quantity_value=shipment.quantity_mt,
            previous_value=None,
            new_value=new_value,
            source_id=event_source_id,
            metadata={"ingestion_file_type": "shipment", "change_type": "created"},
        )
        return ImportRecordChange(
            action="created",
            record_type="shipment",
            record_id=str(shipment.id),
            record_reference=shipment.shipment_id,
            rollback_safe=True,
            previous_value=None,
            new_value=new_value,
        )

    changed_eta_or_state = (
        not values_equal(shipment.current_eta, current_eta) or shipment.current_state != state
    )
    changed = any(not values_equal(getattr(shipment, key), value) for key, value in attrs.items())
    if not changed:
        return ImportRecordChange(
            action="unchanged",
            record_type="shipment",
            record_id=str(shipment.id),
            record_reference=shipment.shipment_id,
            rollback_safe=False,
            previous_value=None,
            new_value=new_value,
        )

    previous_value = shipment_event_value({key: getattr(shipment, key) for key in attrs})
    for key, value in attrs.items():
        setattr(shipment, key, value)
    if changed_eta_or_state:
        db.add(
            ShipmentUpdate(
                tenant_id=context.tenant_id,
                shipment_id=shipment.id,
                source=data["source_of_truth"],
                event_type="shipment_eta_or_state_changed",
                event_time=latest_update_at,
                payload_json=json.dumps(
                    {"current_eta": current_eta.isoformat(), "current_state": state.value}
                ),
                notes="Created by onboarding upload",
            )
        )
        event_type = (
            OperationalEventType.SHIPMENT_ETA_CHANGED
            if not values_equal(previous_value["current_eta"], new_value["current_eta"])
            else OperationalEventType.SHIPMENT_MILESTONE_UPDATED
        )
        emit_shipment_update_event(
            db,
            tenant_id=context.tenant_id,
            event_type=event_type,
            occurred_at=latest_update_at,
            source_reference=source_reference,
            shipment_id=shipment.id,
            shipment_reference=shipment.shipment_id,
            plant_id=plant.id,
            plant_reference=plant.code,
            material_id=material.id,
            material_reference=material.code,
            supplier_id=shipment.supplier_id,
            supplier_reference=shipment.supplier_name,
            quantity_value=shipment.quantity_mt,
            previous_value=previous_value,
            new_value=new_value,
            source_id=event_source_id,
            metadata={"ingestion_file_type": "shipment", "change_type": "updated"},
        )
    return ImportRecordChange(
        action="updated",
        record_type="shipment",
        record_id=str(shipment.id),
        record_reference=shipment.shipment_id,
        rollback_safe=False,
        previous_value=previous_value,
        new_value=new_value,
    )


def upsert_stock_snapshot(
    db: Session,
    context: RequestContext,
    data: dict[str, str],
    event_source_id: int | None = None,
) -> ImportRecordChange:
    plant = resolve_plant(db, context.tenant_id, data["plant_code"])
    material = resolve_material(db, context.tenant_id, data["material_code"])
    daily_consumption = parse_decimal(
        data["daily_consumption_mt"],
        "daily_consumption_mt",
        positive=True,
    )
    snapshot_time = parse_datetime(data["snapshot_time"], "snapshot_time")
    attrs = {
        "on_hand_mt": parse_decimal(data["on_hand_mt"], "on_hand_mt"),
        "quality_held_mt": parse_decimal(data["quality_held_mt"], "quality_held_mt"),
        "available_to_consume_mt": parse_decimal(
            data["available_to_consume_mt"],
            "available_to_consume_mt",
        ),
        "daily_consumption_mt": daily_consumption,
    }
    snapshot = db.scalar(
        select(StockSnapshot).where(
            StockSnapshot.tenant_id == context.tenant_id,
            StockSnapshot.plant_id == plant.id,
            StockSnapshot.material_id == material.id,
            StockSnapshot.snapshot_time == snapshot_time,
        )
    )
    if snapshot is None:
        snapshot = db.scalar(
            select(StockSnapshot)
            .where(
                StockSnapshot.tenant_id == context.tenant_id,
                StockSnapshot.plant_id == plant.id,
                StockSnapshot.material_id == material.id,
            )
            .order_by(StockSnapshot.snapshot_time.desc())
        )
    if snapshot is None:
        snapshot = StockSnapshot(
            tenant_id=context.tenant_id,
            plant_id=plant.id,
            material_id=material.id,
            snapshot_time=snapshot_time,
            **attrs,
        )
        db.add(snapshot)
        db.flush()
        emit_inventory_stock_updated(
            db,
            tenant_id=context.tenant_id,
            occurred_at=snapshot_time,
            source_reference=data.get("source_of_truth"),
            plant_id=plant.id,
            plant_reference=plant.code,
            material_id=material.id,
            material_reference=material.code,
            quantity_value=attrs["available_to_consume_mt"],
            previous_value=None,
            new_value=stock_event_value(attrs, snapshot_time),
            source_id=event_source_id,
            metadata={"ingestion_file_type": "stock", "change_type": "created"},
        )
        return ImportRecordChange(
            action="created",
            record_type="stock_snapshot",
            record_id=str(snapshot.id),
            record_reference=f"{plant.code}/{material.code}/{snapshot_time.isoformat()}",
            rollback_safe=True,
            previous_value=None,
            new_value=stock_event_value(attrs, snapshot_time),
        )

    changed = any(not values_equal(getattr(snapshot, key), value) for key, value in attrs.items())
    if not changed:
        return ImportRecordChange(
            action="unchanged",
            record_type="stock_snapshot",
            record_id=str(snapshot.id),
            record_reference=f"{plant.code}/{material.code}/{snapshot.snapshot_time.isoformat()}",
            rollback_safe=False,
            previous_value=None,
            new_value=stock_event_value(attrs, snapshot.snapshot_time),
        )
    previous_value = stock_event_value(
        {key: getattr(snapshot, key) for key in attrs},
        snapshot.snapshot_time,
    )
    for key, value in attrs.items():
        setattr(snapshot, key, value)
    emit_inventory_stock_updated(
        db,
        tenant_id=context.tenant_id,
        occurred_at=snapshot_time,
        source_reference=data.get("source_of_truth"),
        plant_id=plant.id,
        plant_reference=plant.code,
        material_id=material.id,
        material_reference=material.code,
        quantity_value=attrs["available_to_consume_mt"],
        previous_value=previous_value,
        new_value=stock_event_value(attrs, snapshot_time),
        source_id=event_source_id,
        metadata={"ingestion_file_type": "stock", "change_type": "updated"},
    )
    return ImportRecordChange(
        action="updated",
        record_type="stock_snapshot",
        record_id=str(snapshot.id),
        record_reference=f"{plant.code}/{material.code}/{snapshot.snapshot_time.isoformat()}",
        rollback_safe=False,
        previous_value=previous_value,
        new_value=stock_event_value(
            {
                "on_hand_mt": snapshot.on_hand_mt,
                "quality_held_mt": snapshot.quality_held_mt,
                "available_to_consume_mt": snapshot.available_to_consume_mt,
                "daily_consumption_mt": snapshot.daily_consumption_mt,
            },
            snapshot_time,
        ),
    )


def upsert_consumption(
    db: Session,
    context: RequestContext,
    data: dict[str, str],
    event_source_id: int | None = None,
) -> ImportRecordChange:
    plant = resolve_plant(db, context.tenant_id, data["plant_code"])
    material = resolve_material(db, context.tenant_id, data["material_code"])
    snapshot_time = parse_datetime(data["snapshot_time"], "snapshot_time")
    daily_consumption = parse_decimal(
        data["daily_consumption_mt"],
        "daily_consumption_mt",
        positive=True,
    )
    snapshot = db.scalar(
        select(StockSnapshot).where(
            StockSnapshot.tenant_id == context.tenant_id,
            StockSnapshot.plant_id == plant.id,
            StockSnapshot.material_id == material.id,
            StockSnapshot.snapshot_time == snapshot_time,
        )
    )
    if snapshot is None:
        raise FieldValueError(
            "snapshot_time",
            "Consumption could not be linked to an inventory snapshot for this "
            "plant/material/time.",
            "Load the matching inventory sheet in the same workbook or upload inventory first.",
        )
    if values_equal(snapshot.daily_consumption_mt, daily_consumption):
        return ImportRecordChange(
            action="unchanged",
            record_type="stock_snapshot",
            record_id=str(snapshot.id),
            record_reference=f"{plant.code}/{material.code}/{snapshot.snapshot_time.isoformat()}",
            rollback_safe=False,
            previous_value=None,
            new_value=stock_event_value(
                {
                    "on_hand_mt": snapshot.on_hand_mt,
                    "quality_held_mt": snapshot.quality_held_mt,
                    "available_to_consume_mt": snapshot.available_to_consume_mt,
                    "daily_consumption_mt": snapshot.daily_consumption_mt,
                },
                snapshot.snapshot_time,
            ),
        )
    previous_value = stock_event_value(
        {
            "on_hand_mt": snapshot.on_hand_mt,
            "quality_held_mt": snapshot.quality_held_mt,
            "available_to_consume_mt": snapshot.available_to_consume_mt,
            "daily_consumption_mt": snapshot.daily_consumption_mt,
        },
        snapshot.snapshot_time,
    )
    snapshot.daily_consumption_mt = daily_consumption
    emit_inventory_stock_updated(
        db,
        tenant_id=context.tenant_id,
        occurred_at=snapshot_time,
        source_reference=data.get("source_of_truth"),
        plant_id=plant.id,
        plant_reference=plant.code,
        material_id=material.id,
        material_reference=material.code,
        quantity_value=snapshot.available_to_consume_mt,
        previous_value=previous_value,
        new_value=stock_event_value(
            {
                "on_hand_mt": snapshot.on_hand_mt,
                "quality_held_mt": snapshot.quality_held_mt,
                "available_to_consume_mt": snapshot.available_to_consume_mt,
                "daily_consumption_mt": daily_consumption,
            },
            snapshot_time,
        ),
        source_id=event_source_id,
        metadata={"ingestion_file_type": "consumption", "change_type": "updated"},
    )
    return ImportRecordChange(
        action="updated",
        record_type="stock_snapshot",
        record_id=str(snapshot.id),
        record_reference=f"{plant.code}/{material.code}/{snapshot.snapshot_time.isoformat()}",
        rollback_safe=False,
        previous_value=previous_value,
        new_value=stock_event_value(
            {
                "on_hand_mt": snapshot.on_hand_mt,
                "quality_held_mt": snapshot.quality_held_mt,
                "available_to_consume_mt": snapshot.available_to_consume_mt,
                "daily_consumption_mt": daily_consumption,
            },
            snapshot_time,
        ),
    )


def upsert_threshold(
    db: Session,
    context: RequestContext,
    data: dict[str, str],
) -> ImportRecordChange:
    plant = resolve_plant(db, context.tenant_id, data["plant_code"])
    material = resolve_material(db, context.tenant_id, data["material_code"])
    threshold_days = parse_decimal(data["threshold_days"], "threshold_days", positive=True)
    warning_days = parse_decimal(data["warning_days"], "warning_days", positive=True)
    if warning_days < threshold_days:
        raise FieldValueError(
            "warning_days",
            "Warning days cannot be earlier than the critical threshold.",
            "Set Warning days greater than or equal to Critical threshold days.",
        )
    reserve_quantity = parse_optional_decimal(data, "reserve_quantity_mt")
    quality_hold_quantity = parse_optional_decimal(data, "quality_hold_quantity_mt")
    attrs = {
        "threshold_days": threshold_days,
        "warning_days": warning_days,
        "minimum_buffer_stock_mt": parse_optional_decimal(data, "minimum_buffer_stock_mt"),
        "reserve_quantity_mt": reserve_quantity,
        "quality_hold_quantity_mt": quality_hold_quantity,
        "minimum_buffer_stock_days": parse_decimal(
            data["minimum_buffer_stock_days"],
            "minimum_buffer_stock_days",
        )
        if data.get("minimum_buffer_stock_days")
        else None,
        "stockout_alert_horizon_days": parse_decimal(
            data["stockout_alert_horizon_days"],
            "stockout_alert_horizon_days",
        )
        if data.get("stockout_alert_horizon_days")
        else None,
    }
    threshold = db.scalar(
        select(PlantMaterialThreshold).where(
            PlantMaterialThreshold.tenant_id == context.tenant_id,
            PlantMaterialThreshold.plant_id == plant.id,
            PlantMaterialThreshold.material_id == material.id,
        )
    )
    if threshold is None:
        threshold = PlantMaterialThreshold(
            tenant_id=context.tenant_id,
            plant_id=plant.id,
            material_id=material.id,
            **attrs,
        )
        db.add(threshold)
        db.flush()
        return ImportRecordChange(
            action="created",
            record_type="threshold",
            record_id=str(threshold.id),
            record_reference=f"{plant.code}/{material.code}",
            rollback_safe=True,
            previous_value=None,
            new_value={key: json_value(value) for key, value in attrs.items()},
        )

    changed = any(not values_equal(getattr(threshold, key), value) for key, value in attrs.items())
    if not changed:
        return ImportRecordChange(
            action="unchanged",
            record_type="threshold",
            record_id=str(threshold.id),
            record_reference=f"{plant.code}/{material.code}",
            rollback_safe=False,
            previous_value=None,
            new_value={key: json_value(value) for key, value in attrs.items()},
        )
    previous_value = {
        "threshold_days": json_value(threshold.threshold_days),
        "warning_days": json_value(threshold.warning_days),
        "minimum_buffer_stock_mt": json_value(threshold.minimum_buffer_stock_mt),
        "reserve_quantity_mt": json_value(threshold.reserve_quantity_mt),
        "quality_hold_quantity_mt": json_value(threshold.quality_hold_quantity_mt),
        "minimum_buffer_stock_days": json_value(threshold.minimum_buffer_stock_days),
        "stockout_alert_horizon_days": json_value(threshold.stockout_alert_horizon_days),
    }
    for key, value in attrs.items():
        setattr(threshold, key, value)
    return ImportRecordChange(
        action="updated",
        record_type="threshold",
        record_id=str(threshold.id),
        record_reference=f"{plant.code}/{material.code}",
        rollback_safe=False,
        previous_value=previous_value,
        new_value={key: json_value(value) for key, value in attrs.items()},
    )
