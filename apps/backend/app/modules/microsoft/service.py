from __future__ import annotations

import base64
import io
import json
import logging
import secrets
from datetime import UTC, datetime, timedelta
from pathlib import Path
from typing import Any
from urllib.parse import quote, urlencode

import httpx
from fastapi import HTTPException, status
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.config import settings
from app.models import MicrosoftConnection, MicrosoftDataSource, MicrosoftOAuthState
from app.modules.ingestion.service import process_upload_content
from app.schemas.context import RequestContext
from app.utils.encryption import decrypt, encrypt

logger = logging.getLogger(__name__)

AUTH_ENDPOINT = "https://login.microsoftonline.com/common/oauth2/v2.0/authorize"
TOKEN_ENDPOINT = "https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
GRAPH_BASE = "https://graph.microsoft.com/v1.0"
SCOPES = ["Files.Read", "Files.Read.All", "offline_access", "User.Read"]
SUPPORTED_SUFFIXES = {".csv", ".xls", ".xlsx"}


def get_authorization_url(db: Session, tenant_id: int, user_id: int) -> tuple[str, str]:
    state = secrets.token_hex(16)
    now = datetime.now(UTC)
    db.add(
        MicrosoftOAuthState(
            state=state,
            tenant_id=tenant_id,
            user_id=user_id,
            created_at=now,
            expires_at=now + timedelta(minutes=15),
            used=False,
        )
    )
    db.commit()
    params = {
        "client_id": settings.microsoft_client_id,
        "response_type": "code",
        "redirect_uri": settings.microsoft_redirect_uri,
        "scope": " ".join(SCOPES),
        "state": state,
        "response_mode": "query",
        "prompt": "select_account",
    }
    return f"{AUTH_ENDPOINT}?{urlencode(params)}", state


def handle_callback(db: Session, code: str, state: str) -> MicrosoftConnection:
    oauth_state = _validate_state(db, state)
    oauth_state.used = True
    tokens = _token_post(
        "common",
        {
            "client_id": settings.microsoft_client_id,
            "client_secret": settings.microsoft_client_secret,
            "code": code,
            "redirect_uri": settings.microsoft_redirect_uri,
            "grant_type": "authorization_code",
            "scope": " ".join(SCOPES),
        },
    )
    access_token = tokens["access_token"]
    refresh_token = tokens["refresh_token"]
    claims = _decode_jwt_payload(tokens.get("id_token", ""))
    microsoft_tenant_id = claims.get("tid") or tokens.get("tenant") or "common"
    profile = _graph_json("GET", "/me", access_token)
    microsoft_user_id = str(profile.get("id") or claims.get("oid") or claims.get("sub"))
    email = str(profile.get("mail") or profile.get("userPrincipalName") or "")
    display_name = str(profile.get("displayName") or email or "Microsoft user")
    expires_at = datetime.now(UTC) + timedelta(seconds=int(tokens.get("expires_in", 3600)))

    connection = db.scalar(
        select(MicrosoftConnection).where(
            MicrosoftConnection.tenant_id == oauth_state.tenant_id,
            MicrosoftConnection.microsoft_user_id == microsoft_user_id,
        )
    )
    if connection is None:
        connection = MicrosoftConnection(
            tenant_id=oauth_state.tenant_id,
            microsoft_user_id=microsoft_user_id,
            connected_at=datetime.now(UTC),
        )
        db.add(connection)

    connection.microsoft_tenant_id = str(microsoft_tenant_id)
    connection.display_name = display_name
    connection.email = email
    connection.access_token = encrypt(access_token)
    connection.refresh_token = encrypt(refresh_token)
    connection.token_expires_at = expires_at
    connection.scope = str(tokens.get("scope") or " ".join(SCOPES))
    connection.auth_error = None
    connection.is_active = True
    db.commit()
    db.refresh(connection)
    return connection


def refresh_access_token(db: Session, connection: MicrosoftConnection) -> MicrosoftConnection:
    try:
        tokens = _token_post(
            connection.microsoft_tenant_id or "common",
            {
                "client_id": settings.microsoft_client_id,
                "client_secret": settings.microsoft_client_secret,
                "refresh_token": decrypt(connection.refresh_token),
                "redirect_uri": settings.microsoft_redirect_uri,
                "grant_type": "refresh_token",
                "scope": " ".join(SCOPES),
            },
        )
    except Exception as exc:
        connection.auth_error = str(exc)
        if _is_terminal_auth_error(exc):
            connection.is_active = False
        db.commit()
        return connection

    connection.access_token = encrypt(tokens["access_token"])
    if tokens.get("refresh_token"):
        connection.refresh_token = encrypt(tokens["refresh_token"])
    connection.token_expires_at = datetime.now(UTC) + timedelta(seconds=int(tokens.get("expires_in", 3600)))
    connection.scope = str(tokens.get("scope") or connection.scope)
    connection.last_token_refresh_at = datetime.now(UTC)
    connection.auth_error = None
    connection.is_active = True
    db.commit()
    db.refresh(connection)
    return connection


def get_valid_token(db: Session, connection: MicrosoftConnection) -> str:
    expires_at = _aware(connection.token_expires_at)
    if expires_at > datetime.now(UTC) + timedelta(minutes=5) and connection.is_active:
        return decrypt(connection.access_token)
    refreshed = refresh_access_token(db, connection)
    if refreshed.auth_error or not refreshed.is_active:
        raise PermissionError(refreshed.auth_error or "Microsoft authentication expired")
    return decrypt(refreshed.access_token)


def list_drive_files(
    db: Session,
    connection: MicrosoftConnection,
    drive_id: str | None = None,
    search_query: str | None = None,
) -> list[dict[str, Any]]:
    token = get_valid_token(db, connection)
    items: list[dict[str, Any]] = []
    if search_query:
        items.extend(_collect_drive_search_items(token, search_query, drive_id=drive_id))
    elif drive_id:
        payload = _graph_json("GET", f"/me/drives/{drive_id}/root/children", token)
        items.extend(payload.get("value", []))
        for query in ("stock_snapshot", "stock", "shipments", "threshold", "xlsx", "xls", "csv"):
            items.extend(_collect_drive_search_items(token, query, drive_id=drive_id))
    else:
        payload = _graph_json("GET", "/me/drive/root/children", token)
        items.extend(payload.get("value", []))
        try:
            recent_payload = _graph_json("GET", "/me/drive/recent", token)
            items.extend(recent_payload.get("value", []))
        except Exception:
            logger.debug("Microsoft recent files lookup failed", exc_info=True)
        for query in ("stock_snapshot", "stock", "shipments", "threshold", "xlsx", "xls", "csv"):
            items.extend(_collect_drive_search_items(token, query))
    return _dedupe_files(items, is_sharepoint=False)


def _collect_drive_search_items(
    token: str,
    query: str,
    drive_id: str | None = None,
) -> list[dict[str, Any]]:
    encoded_query = quote(query.replace("'", "''"))
    if drive_id:
        paths = [f"/me/drives/{drive_id}/root/search(q='{encoded_query}')"]
    else:
        paths = [
            f"/me/drive/root/search(q='{encoded_query}')",
            f"/me/drive/search(q='{encoded_query}')",
        ]
    items: list[dict[str, Any]] = []
    for path in paths:
        try:
            payload = _graph_json("GET", path, token)
            items.extend(payload.get("value", []))
        except Exception:
            logger.debug("Microsoft file search failed", extra={"path": path}, exc_info=True)
    return items


def list_sharepoint_sites(db: Session, connection: MicrosoftConnection) -> list[dict[str, str]]:
    token = get_valid_token(db, connection)
    payload = _graph_json("GET", "/sites?search=*", token)
    return [
        {
            "site_id": item["id"],
            "name": item.get("name") or item.get("displayName") or "SharePoint site",
            "display_name": item.get("displayName") or item.get("name") or "SharePoint site",
        }
        for item in payload.get("value", [])
    ]


def list_site_drives(db: Session, connection: MicrosoftConnection, site_id: str) -> list[dict[str, str]]:
    token = get_valid_token(db, connection)
    payload = _graph_json("GET", f"/sites/{site_id}/drives", token)
    return [{"drive_id": item["id"], "name": item.get("name") or "Documents"} for item in payload.get("value", [])]


def list_sharepoint_files(
    db: Session,
    connection: MicrosoftConnection,
    site_id: str,
    drive_id: str,
) -> list[dict[str, Any]]:
    token = get_valid_token(db, connection)
    payload = _graph_json("GET", f"/sites/{site_id}/drives/{drive_id}/root/children", token)
    return [_file_payload(item, is_sharepoint=True) for item in payload.get("value", []) if _is_supported_file(item)]


def download_file(
    db: Session,
    connection: MicrosoftConnection,
    drive_id: str,
    item_id: str,
    site_id: str | None = None,
) -> bytes:
    path = (
        f"/sites/{site_id}/drives/{drive_id}/items/{item_id}/content"
        if site_id
        else f"/me/drives/{drive_id}/items/{item_id}/content"
    )
    token = get_valid_token(db, connection)
    response = _graph_response("GET", path, token, follow_redirects=True)
    if response.status_code == status.HTTP_401_UNAUTHORIZED:
        token = decrypt(refresh_access_token(db, connection).access_token)
        response = _graph_response("GET", path, token, follow_redirects=True)
    if response.status_code >= 400:
        raise RuntimeError(f"Microsoft Graph download failed ({response.status_code})")
    return response.content


def get_file_metadata(
    db: Session,
    connection: MicrosoftConnection,
    drive_id: str,
    item_id: str,
    site_id: str | None = None,
) -> dict[str, Any]:
    token = get_valid_token(db, connection)
    path = (
        f"/sites/{site_id}/drives/{drive_id}/items/{item_id}"
        if site_id
        else f"/me/drives/{drive_id}/items/{item_id}"
    )
    return _graph_json("GET", path, token)


def get_sheet_names(
    db: Session,
    connection: MicrosoftConnection,
    drive_id: str,
    item_id: str,
    site_id: str | None = None,
) -> list[str]:
    content = download_file(db, connection, drive_id, item_id, site_id=site_id)
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    return list(workbook.sheetnames)


def sync_microsoft_data_source(db: Session, source: MicrosoftDataSource) -> dict[str, Any]:
    source.sync_status = "syncing"
    source.last_sync_attempted_at = datetime.now(UTC)
    source.last_sync_error = None
    db.commit()
    connection = db.get(MicrosoftConnection, source.microsoft_connection_id)
    if connection is None or connection.tenant_id != source.tenant_id or not connection.is_active:
        source.sync_status = "auth_error"
        source.last_sync_error = "Microsoft connection is not active"
        db.commit()
        raise PermissionError(source.last_sync_error)

    try:
        metadata = get_file_metadata(db, connection, source.drive_id, source.item_id, source.site_id)
        filename = metadata.get("name") or source.display_name or "microsoft-source.xlsx"
        content = download_file(db, connection, source.drive_id, source.item_id, source.site_id)
        if source.sheet_name and Path(filename).suffix.lower() == ".xlsx":
            content = activate_sheet(content, source.sheet_name)
        result = process_upload_content(
            db=db,
            context=RequestContext(
                tenant_id=source.tenant_id,
                tenant_slug="background",
                role="tenant_admin",
                user_id=0,
            ),
            current_user_id=None,
            file_type=source.file_type,
            filename=filename,
            content=content,
            content_type=_content_type(filename),
            mapping_overrides=source.column_mapping,
            source_of_truth="microsoft_graph",
        )
        db.add(source)
        source.sync_status = "success"
        source.last_successful_sync_at = datetime.now(UTC)
        source.last_sync_error = None
        db.commit()
        return {
            "rows_ingested": result.rows_accepted,
            "file_type": source.file_type,
            "status": "success",
            "detail": result.model_dump(),
        }
    except PermissionError as exc:
        db.add(source)
        source.sync_status = "auth_error"
        source.last_sync_error = str(exc)
        db.commit()
        raise
    except Exception as exc:
        db.add(source)
        source.sync_status = "error"
        source.last_sync_error = str(exc)
        db.commit()
        raise


def _validate_state(db: Session, state: str) -> MicrosoftOAuthState:
    oauth_state = db.scalar(select(MicrosoftOAuthState).where(MicrosoftOAuthState.state == state))
    if oauth_state is None or oauth_state.used or _aware(oauth_state.expires_at) < datetime.now(UTC):
        raise HTTPException(status_code=400, detail="Invalid or expired Microsoft OAuth state")
    return oauth_state


def _token_post(tenant_id: str, data: dict[str, str]) -> dict[str, Any]:
    response = httpx.post(TOKEN_ENDPOINT.format(tenant_id=tenant_id), data=data, timeout=20)
    if response.status_code >= 400:
        try:
            body = response.json()
        except ValueError:
            body = {}
        error_code = body.get("error")
        description = body.get("error_description", "Microsoft token exchange failed")
        raise RuntimeError(f"{error_code or 'token_error'}: {description}")
    return response.json()


def _is_terminal_auth_error(exc: Exception) -> bool:
    message = str(exc).lower()
    return "invalid_grant" in message or "interaction_required" in message


def _graph_json(method: str, path: str, token: str) -> dict[str, Any]:
    response = _graph_response(method, path, token)
    if response.status_code >= 400:
        raise RuntimeError(f"Microsoft Graph request failed ({response.status_code})")
    return response.json()


def _graph_response(method: str, path: str, token: str, follow_redirects: bool = False) -> httpx.Response:
    url = path if path.startswith("http") else f"{GRAPH_BASE}{path}"
    return httpx.request(
        method,
        url,
        headers={"Authorization": f"Bearer {token}"},
        timeout=30,
        follow_redirects=follow_redirects,
    )


def _decode_jwt_payload(token: str) -> dict[str, Any]:
    if not token or token.count(".") < 2:
        return {}
    payload = token.split(".")[1]
    payload += "=" * (-len(payload) % 4)
    try:
        return json.loads(base64.urlsafe_b64decode(payload.encode()))
    except Exception:
        return {}


def _file_payload(item: dict[str, Any], is_sharepoint: bool) -> dict[str, Any]:
    drive_id = item.get("parentReference", {}).get("driveId") or item.get("driveId") or ""
    return {
        "item_id": item["id"],
        "drive_id": drive_id,
        "name": item.get("name") or "Untitled file",
        "size": item.get("size"),
        "modified_at": item.get("lastModifiedDateTime"),
        "web_url": item.get("webUrl"),
        "is_sharepoint": is_sharepoint,
    }


def _dedupe_files(items: list[dict[str, Any]], is_sharepoint: bool) -> list[dict[str, Any]]:
    seen: set[tuple[str, str]] = set()
    files: list[dict[str, Any]] = []
    for item in items:
        if not _is_supported_file(item):
            continue
        payload = _file_payload(item, is_sharepoint=is_sharepoint)
        key = (payload["drive_id"], payload["item_id"])
        if key in seen:
            continue
        seen.add(key)
        files.append(payload)
    return files


def _is_supported_file(item: dict[str, Any]) -> bool:
    if "file" not in item:
        return False
    return Path(str(item.get("name") or "")).suffix.lower() in SUPPORTED_SUFFIXES


def activate_sheet(content: bytes, sheet_name: str) -> bytes:
    workbook = load_workbook(io.BytesIO(content))
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' was not found")
    workbook.active = workbook.sheetnames.index(sheet_name)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _content_type(filename: str) -> str:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        return "text/csv"
    if suffix == ".xls":
        return "application/vnd.ms-excel"
    return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _aware(value: datetime) -> datetime:
    if value.tzinfo is None:
        return value.replace(tzinfo=UTC)
    return value
